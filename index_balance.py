#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Index color-balance checker & lane balancer for Illumina 2-color (NovaSeq X/X Plus).

Rules:
  T: green only
  C: green + blue
  A: blue only
  G: dark

Hard constraints (strict):
  - Cycle 1 and 2 must satisfy channel thresholds (per index type i7/i5).

Weighted optimization (soft, for cycles >=3):
  - By default, cycles 3/4 share 60% of the total weight (evenly),
    and cycles 5..N share the remaining 40% (evenly).
  - Cycles 1 and 2 have zero weight because they are strictly enforced.
  - You can adjust the 60% via --w34 (0~1).

Input CSV/TSV must have headers: id,i7,i5
i5 is NOT reverse-complemented by default. Use --rc-i5 if needed.

easy run:
  python index_balance.py group library.csv --lanes 3 --out lanes.csv 
  # library.csv is a csv file with headers: id,i7,i5
  # xlsx is also supported

  python index_balance.py check lanes.csv 
  # lanes.csv is a csv file with headers: id,i7,i5
  # xlsx is also supported

Outputs (human-readable):
  - check: overall info + per-cycle color fractions and pass/fail for i7 and i5
  - group: lane summary + per-cycle details for each lane (i7 and i5)
"""

import argparse
import csv
import random
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import openpyxl

BASES = ("A", "C", "G", "T")
BIG_PENALTY = 1e9  # for violating hard constraints in optimization

# -------------------- IO & utils --------------------

def revcomp(seq: str) -> str:
    comp = str.maketrans("ACGTNacgtn", "TGCANtgcan")
    return seq.translate(comp)[::-1]

def read_table(path: str) -> List[Dict[str, str]]:
    """
    Read CSV/TSV or XLSX with headers id,i7,i5; handle CSV BOM; trim header whitespace.
    - CSV/TSV: auto delimiter by suffix (.csv => ',', otherwise '\t'), encoding='utf-8-sig'
    - XLSX/XLSM: use openpyxl, read the first sheet's first non-empty row as headers
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    suffix = p.suffix.lower()
    items: List[Dict[str, str]] = []

    # ---- Excel path ----
    if suffix in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise ImportError(
                "Reading XLSX/XLSM requires 'openpyxl'. Install with: pip install openpyxl"
            )
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb.active  # 使用第一个工作表

        # 找到第一行非空作为表头
        header = None
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            vals = [("" if v is None else str(v)).strip() for v in row]
            if any(v != "" for v in vals):
                header = vals
                break
        if not header:
            raise ValueError("No header row found in the Excel sheet.")

        # 规范化表头（去空白、去BOM、小写）
        header = [h.replace("\ufeff", "").strip() for h in header]
        header_lower = [h.lower() for h in header]
        # 建立列名->索引映射
        colmap = {name: idx for idx, name in enumerate(header_lower)}

        required = {"id", "i7", "i5"}
        missing = required - set(colmap.keys())
        if missing:
            raise ValueError(f"Missing required columns: {missing}. Found: {header}")

        # 继续读数据行
        for row in ws.iter_rows(min_row=ws.min_row + 1, values_only=True):
            if row is None:
                continue
            def get(col):
                idx = colmap[col]
                v = "" if idx >= len(row) or row[idx] is None else str(row[idx])
                return v.strip()
            rec = {"id": get("id"), "i7": get("i7").upper(), "i5": get("i5").upper()}
            # 跳过完全空行
            if not (rec["id"] or rec["i7"] or rec["i5"]):
                continue
            items.append(rec)
        wb.close()
        return items

    # ---- CSV/TSV path ----
    # 默认：.csv 用逗号，其他后缀（.tsv/.txt）用制表符
    delim = "," if suffix == ".csv" else "\t"
    with p.open(encoding="utf-8-sig") as f:  # 处理 BOM
        reader = csv.DictReader(f, delimiter=delim)
        # 规范化表头
        fieldnames = [fn.replace("\ufeff", "").strip() for fn in (reader.fieldnames or [])]
        reader.fieldnames = fieldnames
        required = {"id", "i7", "i5"}
        missing = required - set(fn.lower() for fn in fieldnames)
        if missing:
            raise ValueError(f"Missing required columns: {missing}. Found: {fieldnames}")
        for row in reader:
            items.append({
                "id": (row.get("id") or row.get("ID") or "").strip(),
                "i7": (row.get("i7") or row.get("I7") or "").strip().upper(),
                "i5": (row.get("i5") or row.get("I5") or "").strip().upper()
            })
    return items

def validate_indexes(items: List[Dict[str, str]]) -> None:
    for r in items:
        for k in ("i7", "i5"):
            s = r[k]
            if not s:
                raise ValueError(f"{r['id']} has empty {k}.")
            for ch in s:
                if ch not in "ACGTN":
                    raise ValueError(f"{r['id']} {k} contains invalid base '{ch}'. Only A/C/G/T/N allowed.")

def effective_cycles(items: List[Dict[str, str]], which: str, requested: int) -> int:
    """Use min(requested, min_length_across_items)."""
    min_len = min(len(r[which]) for r in items) if items else 0
    return max(0, min(requested, min_len))

# -------------------- Color math --------------------

def per_cycle_base_lists(seqs: List[str], L: int) -> List[List[str]]:
    """Collect bases per cycle (ignore N)."""
    cycles = [[] for _ in range(L)]
    for s in seqs:
        for i in range(L):
            b = s[i]
            if b in BASES:
                cycles[i].append(b)
    return cycles

def color_fractions_one_cycle(base_list: List[str]) -> Dict[str, float]:
    """
    2-color mapping:
      T: green only
      C: green + blue
      A: blue only
      G: dark
    Each base contributes up to 2 "signal units".
    """
    a = base_list.count('A')
    t = base_list.count('T')
    c = base_list.count('C')
    g = base_list.count('G')
    tot_units = 2 * (a + t + c + g)
    if tot_units == 0:
        return {"green": 0.0, "blue": 0.0, "dark": 0.0}
    green = (t * 2.0 + c * 1.0) / tot_units
    blue  = (a * 2.0 + c * 1.0) / tot_units
    dark  = (g * 2.0) / tot_units
    return {"green": green, "blue": blue, "dark": dark}

def hard_cycles_ok_aggregate(
    seqs: List[str],
    L: int,
    min_green: float,
    min_blue: float,
    max_dark: float
) -> Tuple[bool, List[Optional[Dict[str, float]]]]:
    """
    Check hard constraints for cycle 1 and 2 on the aggregate (lane-level / batch-level).
    Returns (ok, [c1_color, c2_color]) where entries may be None if L<cycle.
    """
    blists = per_cycle_base_lists(seqs, L)
    colors_12: List[Optional[Dict[str, float]]] = [None, None]
    for idx in (0, 1):  # cycles 1 and 2
        if L > idx:
            col = color_fractions_one_cycle(blists[idx])
            colors_12[idx] = col
            ok = (col["green"] >= min_green) and (col["blue"] > min_blue) and (col["dark"] <= max_dark)
            if not ok:
                return False, colors_12
    return True, colors_12

def make_weights_special(L: int, w34: float) -> List[float]:
    """
    Build per-cycle weights for penalty (sum to 1 over cycles >=3; cycles 1-2 weight=0).
    - If L <= 2: all zeros.
    - If 3 <= L <= 4: distribute all weight uniformly over cycles 3..L.
    - If L > 4: cycles 3/4 share 'w34' (evenly), cycles 5..L share (1-w34) (evenly).
    """
    if L <= 2:
        return [0.0] * L

    # clamp w34
    if w34 < 0.0:
        w34 = 0.0
    if w34 > 1.0:
        w34 = 1.0

    w = [0.0] * L
    if 3 <= L <= 4:
        # all mass spread over 3..L
        mass = 1.0
        n = L - 2  # number of cycles starting from 3
        for i in range(2, L):
            w[i] = mass / n
    else:
        # L > 4
        # cycles 3 and 4
        w[2] = w34 / 2.0  # cycle 3
        w[3] = w34 / 2.0  # cycle 4
        # cycles 5..L
        n_tail = L - 4
        if n_tail > 0:
            tail_share = (1.0 - w34) / n_tail
            for i in range(4, L):
                w[i] = tail_share

    # numerical safety: normalize to sum 1 over cycles >=3
    total = sum(w)
    if total > 0:
        w = [x / total for x in w]
    return w

def weighted_penalty(
    seqs: List[str],
    L: int,
    min_green: float,
    min_blue: float,
    max_dark: float,
    weights: List[float]
) -> float:
    """Sum_i weights[i]*vi, where vi is threshold violation at cycle i (cycles 1-2 weights should be 0)."""
    if L <= 0 or not seqs or not weights:
        return 0.0
    blists = per_cycle_base_lists(seqs, L)
    pen = 0.0
    for i in range(min(L, len(weights))):
        if weights[i] == 0.0:
            continue
        col = color_fractions_one_cycle(blists[i])
        v = 0.0
        v += max(0.0, (min_green - col["green"]))
        v += max(0.0, (min_blue  - col["blue"]))
        v += max(0.0, (col["dark"] - max_dark))
        pen += weights[i] * v
    return pen

def evaluate_color_balance_strict_weighted(
    seqs: List[str],
    cycles: int,
    min_green: float,
    min_blue: float,
    max_dark: float,
    w34: float
) -> Dict[str, Any]:
    """
    For CHECK / per-lane report:
      - Enforce hard constraints on cycles 1 and 2.
      - Report per-cycle pass/fail for 1..cycles.
      - Report weighted penalty over cycles >=3 using make_weights_special(cycles, w34).
    """
    # hard constraints on C1/C2
    hard_ok, colors_12 = hard_cycles_ok_aggregate(seqs, cycles, min_green, min_blue, max_dark)

    # per-cycle info
    blists = per_cycle_base_lists(seqs, cycles)
    per_cycle: List[Dict[str, Any]] = []
    worst = None
    worst_score = -1.0
    for i, bl in enumerate(blists, start=1):
        col = color_fractions_one_cycle(bl)
        ok = (col["green"] >= min_green) and (col["blue"] > min_blue) and (col["dark"] <= max_dark)
        per_cycle.append({"cycle": i, "color": col, "ok": ok})
        # total violation for "worst cycle"
        v = max(0.0, (min_green - col["green"])) \
          + max(0.0, (min_blue  - col["blue"])) \
          + max(0.0, (col["dark"] - max_dark))
        if v > worst_score:
            worst_score = v
            worst = {"cycle": i, "color": col, "color_ok": (v == 0.0)}

    # mean color (informational)
    denom = cycles if cycles > 0 else 1
    mean_color = {
        "green": (sum(pc["color"]["green"] for pc in per_cycle) / denom) if per_cycle else 0.0,
        "blue":  (sum(pc["color"]["blue"]  for pc in per_cycle) / denom) if per_cycle else 0.0,
        "dark":  (sum(pc["color"]["dark"]  for pc in per_cycle) / denom) if per_cycle else 0.0
    }

    # weighted penalty over cycles >=3
    weights = make_weights_special(cycles, w34)
    wpen = weighted_penalty(seqs, cycles, min_green, min_blue, max_dark, weights)

    return {
        "cycles": cycles,
        "hard_ok": hard_ok,
        "hard_c12": colors_12,           # [c1_color, c2_color] or None
        "per_cycle": per_cycle,          # list of {cycle, color{...}, ok}
        "weighted_penalty": wpen,        # cycles >=3
        "weights": weights,              # transparency
        "mean_color": mean_color,
        "worst_cycle": worst
    }

# -------------------- Lane optimization (with hard constraints) --------------------

def lane_cost_with_hard(
    i7_seqs: List[str],
    i5_seqs: List[str],
    L7: int, L5: int,
    min_green: float, min_blue: float, max_dark: float,
    w7: List[float],
    w5: List[float]
) -> float:
    """Cost = BIG_PENALTY if any hard constraint fails; else sum of weighted penalties (i7+i5)."""
    ok7, _ = hard_cycles_ok_aggregate(i7_seqs, L7, min_green, min_blue, max_dark)
    ok5, _ = hard_cycles_ok_aggregate(i5_seqs, L5, min_green, min_blue, max_dark)
    if not ok7 or not ok5:
        return BIG_PENALTY
    pen7 = weighted_penalty(i7_seqs, L7, min_green, min_blue, max_dark, w7)
    pen5 = weighted_penalty(i5_seqs, L5, min_green, min_blue, max_dark, w5)
    return pen7 + pen5

def initial_round_robin(items: List[Dict[str, str]], lanes: int) -> List[List[Dict[str, str]]]:
    buckets = [[] for _ in range(lanes)]
    keyed = sorted(items, key=lambda r: (r["i7"] + "_" + r["i5"]))
    for i, r in enumerate(keyed):
        buckets[i % lanes].append(r)
    return buckets

def optimize_assignment(
    buckets: List[List[Dict[str, str]]],
    L7: int, L5: int,
    rc_i5: bool,
    iters: int,
    seed: int,
    min_green: float, min_blue: float, max_dark: float,
    w7: List[float],
    w5: List[float]
) -> List[List[Dict[str, str]]]:
    random.seed(seed)
    lanes = len(buckets)

    def cost(buck: List[Dict[str, str]]) -> float:
        i7 = [r["i7"][:L7] for r in buck]
        i5s = [revcomp(r["i5"]) if rc_i5 else r["i5"] for r in buck]
        i5 = [s[:L5] for s in i5s]
        return lane_cost_with_hard(i7, i5, L7, L5, min_green, min_blue, max_dark, w7, w5)

    lane_scores = [cost(b) for b in buckets]
    total = sum(lane_scores)

    for _ in range(iters):
        if lanes < 2:
            break
        a, b = random.sample(range(lanes), 2)
        if not buckets[a] or not buckets[b]:
            continue
        ia = random.randrange(len(buckets[a]))
        ib = random.randrange(len(buckets[b]))

        ra = buckets[a][ia]
        rb = buckets[b][ib]
        buckets[a][ia], buckets[b][ib] = rb, ra

        new_a = cost(buckets[a])
        new_b = cost(buckets[b])
        new_total = total - lane_scores[a] - lane_scores[b] + new_a + new_b

        if new_total < total:
            lane_scores[a], lane_scores[b] = new_a, new_b
            total = new_total
        else:
            buckets[a][ia], buckets[b][ib] = ra, rb

    return buckets

# -------------------- Pretty printing --------------------

def fmt_bool(b: bool) -> str:
    return "PASS" if b else "FAIL"

def fmt_color_triplet(d: Dict[str, float]) -> str:
    return f"green={d['green']:.2f}  blue={d['blue']:.2f}  dark={d['dark']:.2f}"

def print_per_cycle_list(per_cycle: List[Dict[str, Any]], indent: str = "  ") -> None:
    print(f"{indent}Per-cycle results:")
    for pc in per_cycle:
        print(f"{indent}  #{pc['cycle']:>2}  {fmt_color_triplet(pc['color'])}  [{'PASS' if pc['ok'] else 'FAIL'}]")

def print_check_human(tag: str, rep: Dict[str, Any], min_green: float, min_blue: float, max_dark: float) -> None:
    print(f"{tag}: {fmt_bool(rep['hard_ok'])}  (HARD: cycles 1-2 must pass)")
    # c1/c2 details
    c1, c2 = rep["hard_c12"]
    if c1 is not None:
        ok1 = (c1["green"] >= min_green) and (c1["blue"] > min_blue) and (c1["dark"] <= max_dark)
        print(f"  C1: {fmt_color_triplet(c1)}  [{'PASS' if ok1 else 'FAIL'}]")
    else:
        print(f"  C1: N/A")
    if c2 is not None:
        ok2 = (c2["green"] >= min_green) and (c2["blue"] > min_blue) and (c2["dark"] <= max_dark)
        print(f"  C2: {fmt_color_triplet(c2)}  [{'PASS' if ok2 else 'FAIL'}]")
    else:
        print(f"  C2: N/A")

    print(f"  Mean color (1..{rep['cycles']}): {fmt_color_triplet(rep['mean_color'])}")
    wc = rep['worst_cycle']
    if wc:
        print(f"  Worst cycle: #{wc['cycle']}  {fmt_color_triplet(wc['color'])}  [{'PASS' if wc['color_ok'] else 'FAIL'}]")
    # weighted penalty info
    weights_info = rep["weights"]
    if any(w > 0 for w in weights_info):
        print(f"  Weighted penalty (cycles ≥3): {rep['weighted_penalty']:.4f}  "
              f"(w34={sum(weights_info[2:4]):.2f} over C3-4)")
    # per-cycle dump
    print_per_cycle_list(rep["per_cycle"], indent="  ")
    print()

def print_group_human(report: Dict[str, Any], out_path: str = "") -> None:
    print("=== Lane Assignment (Color Balance) ===")
    total_libs = sum(l['n_libraries'] for l in report['lane_reports'])
    print(f"Total libraries: {total_libs}")
    print(f"Cycles used: i7={report['cycles']['i7']}, i5={report['cycles']['i5']}")
    print(f"Number of lanes: {report['lanes']}\n")

    for lane in report["lane_reports"]:
        status = "PASS" if (lane["i7"]["hard_ok"] and lane["i5"]["hard_ok"]) else "FAIL"
        print(f"Lane {lane['lane']}: {status}  ({lane['n_libraries']} libraries)")
        print(f"  i7 mean: {fmt_color_triplet(lane['i7']['mean_color'])}")
        print(f"  i5 mean: {fmt_color_triplet(lane['i5']['mean_color'])}")
        print(f"  i7 weighted penalty (≥3): {lane['i7']['weighted_penalty']:.4f}")
        print(f"  i5 weighted penalty (≥3): {lane['i5']['weighted_penalty']:.4f}")
        # Per-cycle details for i7 and i5
        print("  i7:")
        print_per_cycle_list(lane["i7"]["per_cycle"], indent="    ")
        print("  i5:")
        print_per_cycle_list(lane["i5"]["per_cycle"], indent="    ")
        print()

    print(f"All lanes pass (hard C1-2): {'YES' if report['all_lanes_hard_pass'] else 'NO'}")
    if out_path:
        print(f"Assignment saved to: {out_path}")

# -------------------- Commands --------------------

def cmd_check(args):
    items = read_table(args.input)
    validate_indexes(items)

    # cycles (clip to available)
    L7_req = args.i7_cycles if args.i7_cycles is not None else args.cycles
    L5_req = args.i5_cycles if args.i5_cycles is not None else args.cycles
    L7 = effective_cycles(items, "i7", L7_req)
    L5 = effective_cycles(items, "i5", L5_req)

    # sequences (i5 NOT RC by default)
    i7_seqs = [r["i7"][:L7] for r in items]
    i5_full = [revcomp(r["i5"]) if args.rc_i5 else r["i5"] for r in items]
    i5_seqs = [s[:L5] for s in i5_full]

    rep7 = evaluate_color_balance_strict_weighted(
        i7_seqs, L7, args.min_green, args.min_blue, args.max_dark, args.w34
    )
    rep5 = evaluate_color_balance_strict_weighted(
        i5_seqs, L5, args.min_green, args.min_blue, args.max_dark, args.w34
    )

    overall_pass = rep7["hard_ok"] and rep5["hard_ok"]

    print("=== Index Color Balance Check ===")
    print(f"Total libraries: {len(items)}")
    print(f"Cycles used: i7={L7}, i5={L5}\n")
    print_check_human("i7", rep7, args.min_green, args.min_blue, args.max_dark)
    print_check_human("i5", rep5, args.min_green, args.min_blue, args.max_dark)
    print(f"Overall (hard C1-2): {'PASS' if overall_pass else 'FAIL'}")

def cmd_group(args):
    items = read_table(args.input)
    validate_indexes(items)
    if args.lanes < 1:
        raise ValueError("--lanes must be >= 1")

    L7_req = args.i7_cycles if args.i7_cycles is not None else args.cycles
    L5_req = args.i5_cycles if args.i5_cycles is not None else args.cycles
    L7 = effective_cycles(items, "i7", L7_req)
    L5 = effective_cycles(items, "i5", L5_req)

    # weights for optimization
    w7 = make_weights_special(L7, args.w34)
    w5 = make_weights_special(L5, args.w34)

    # initial buckets + optimize (hard C1-2 + weighted ≥3)
    buckets = initial_round_robin(items, args.lanes)
    buckets = optimize_assignment(
        buckets, L7=L7, L5=L5, rc_i5=args.rc_i5,
        iters=args.iters, seed=args.seed,
        min_green=args.min_green, min_blue=args.min_blue, max_dark=args.max_dark,
        w7=w7, w5=w5
    )

    # per-lane reports + optional CSV
    out_rows: List[Dict[str, Any]] = []
    lane_reports: List[Dict[str, Any]] = []
    all_hard_pass = True

    for lane_id, buck in enumerate(buckets, start=1):
        i7 = [r["i7"][:L7] for r in buck]
        i5raw = [r["i5"] for r in buck]
        i5 = [(revcomp(s) if args.rc_i5 else s)[:L5] for s in i5raw]

        rep7 = evaluate_color_balance_strict_weighted(i7, L7, args.min_green, args.min_blue, args.max_dark, args.w34)
        rep5 = evaluate_color_balance_strict_weighted(i5, L5, args.min_green, args.min_blue, args.max_dark, args.w34)

        lane_reports.append({
            "lane": lane_id,
            "n_libraries": len(buck),
            "i7": rep7,
            "i5": rep5
        })

        all_hard_pass = all_hard_pass and rep7["hard_ok"] and rep5["hard_ok"]

        for r in buck:
            out_rows.append({"lane": lane_id, "id": r["id"], "i7": r["i7"], "i5": r["i5"]})

    report = {
        "lanes": args.lanes,
        "cycles": {"i7": L7, "i5": L5},
        "lane_reports": lane_reports,
        "all_lanes_hard_pass": all_hard_pass
    }

    print_group_human(report, out_path=args.out)

    # write assignment
    if args.out:
        p = Path(args.out)
        p.parent.mkdir(parents=True, exist_ok=True)
        with p.open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["lane", "id", "i7", "i5"])
            w.writeheader()
            for row in out_rows:
                w.writerow(row)

# -------------------- CLI --------------------

def build_parser():
    p = argparse.ArgumentParser(
        description="Index color-balance checker and lane balancer (2-color; hard C1-2; weighted ≥3; per-cycle outputs)."
    )
    sub = p.add_subparsers(dest="cmd", required=True)

    def add_common(a):
        a.add_argument("input", help="Input CSV/TSV/XLSX with columns: id,i7,i5")
        # cycles control
        a.add_argument("--cycles", type=int, default=8,
                       help="Number of cycles to evaluate for BOTH i7 and i5 (default 8)")
        a.add_argument("--i7-cycles", type=int, default=None,
                       help="Override number of cycles for i7 (default: use --cycles)")
        a.add_argument("--i5-cycles", type=int, default=None,
                       help="Override number of cycles for i5 (default: use --cycles)")
        # color thresholds
        a.add_argument("--min-green", type=float, default=0.25,
                       help="Min green fraction per cycle (default 0.25)")
        a.add_argument("--min-blue", type=float, default=0.05,
                       help="Min blue fraction per cycle (default 0.05)")
        a.add_argument("--max-dark", type=float, default=0.40,
                       help="Max dark fraction per cycle (default 0.40)")
        # weighting
        a.add_argument("--w34", type=float, default=0.60,
                       help="Total weight assigned to cycles 3/4 (even split). Remaining weight goes to 5..N (default 0.60).")
        # orientation
        a.add_argument("--rc-i5", action="store_true",
                       help="Reverse-complement i5 before evaluating (default: OFF)")

    # check
    pc = sub.add_parser("check", help="Check color balance with hard C1-2 and weighted ≥3 cycles; print per-cycle results.")
    add_common(pc)
    pc.set_defaults(func=cmd_check)

    # group
    pg = sub.add_parser("group", help="Assign libraries into lanes (hard C1-2 + weighted ≥3); print per-cycle results per lane.")
    add_common(pg)
    pg.add_argument("--lanes", type=int, required=True, help="Number of lanes")
    pg.add_argument("--iters", type=int, default=20000,
                    help="Swap iterations for optimization (default 20000)")
    pg.add_argument("--seed", type=int, default=42, help="Random seed (default 42)")
    pg.add_argument("--out", type=str, default="",
                    help="Write lane assignment to CSV (columns: lane,id,i7,i5)")
    pg.set_defaults(func=cmd_group)

    return p

def main():
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)

if __name__ == "__main__":
    main()
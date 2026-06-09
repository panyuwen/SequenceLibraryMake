#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Index color-balance checker & lane balancer for Illumina 2-color (NovaSeq X/X Plus) — v3.

What's new in v3 (vs v2)
========================
1) AUTOMATIC duplicate-index separation (no manual specification).
   - Samples that share the SAME (i7,i5) barcode CANNOT be demultiplexed if placed
     in the same lane. v3 detects every duplicated barcode up front and treats
     "two samples with the same barcode must go to different lanes" as a HARD
     constraint. It builds a conflict graph and produces a conflict-free assignment
     by construction (graph coloring), then optimizes color balance on top.

2) ATOMIC pooling units / BUNDLES (samples that must stay together).
   - Pre-pooled libraries (e.g. exome captures pooled before sequencing) cannot be
     split. Add a `bundle` column: rows sharing the same non-empty bundle id are an
     atomic UNIT and are always assigned to the same lane.
   - A bundle whose own members collide on a barcode is flagged as an UNFIXABLE
     internal conflict (no lane assignment can demultiplex it).

3) FEASIBILITY-FIRST analysis for a given number of lanes.
   - If the libraries CAN be placed into the requested --lanes with no barcode
     conflicts (respecting bundles), v3 gives an optimized color-balanced assignment
     (best found by local search, not guaranteed globally optimal; same spirit as v2).
   - If they CANNOT:
       (a) it reports the MINIMUM number of lanes actually required, and
       (b) without adding lanes, it proposes one or more REMOVAL PLANS — which
           samples to drop to fit into --lanes — and for each plan prints the
           per-lane composition and the per-lane ratio-sum balance.
     Use --auto-expand to instead automatically re-run at the minimum feasible
     number of lanes.

Color rules (Illumina 2-color), unchanged from v2:
  T: green only
  C: green + blue
  A: blue only
  G: dark

Feasibility-first color balance (cycles 1/2), unchanged from v2:
  - The optimizer prefers any assignment whose every lane meets the cycle-1/2
    thresholds (per i7/i5, lane-aggregate). If impossible, it minimizes total
    deviation on cycles 1/2 (lexicographic), while still optimizing cycles >=3.

Weighted optimization (soft, cycles >=3), unchanged from v2:
  - Cycles 3/4 share --w34 of the weight (even split); cycles 5..N share the rest.

Freeze mode (optional): --into 1,2,3
  - Honors the input `lane` column: rows that already have a lane are FIXED there; blank-lane
    rows are distributed into the --into lanes; rows pinned to lanes NOT in --into are ignored
    (e.g. already-finalized lanes). Rows with no/partial index act as ratio-only BALLAST: they
    count toward a lane's read budget (ratio balance) but carry no barcode, so they are excluded
    from conflict checking and color scoring. Color (C1/C2 + >=3) is optimized ONLY for lanes
    whose members are all fully indexed; a lane containing a pool/ballast is left to the pool's
    own complexity. This lets the whole sheet (pre-pinned pools + frozen lanes + blanks) be
    solved in one command. BALLAST is unit-level: a `bundle` must be all-indexed or all
    no-index (a mixed bundle is rejected). --into and --constraints are mutually exclusive.

Manual constraints (optional, highest priority): --constraints <json|yaml>
  - Per-lane 'quotas' (type -> sample count) and 'must_include' (sample id or bundle id
    pinned to a lane). When given, --constraints sets the lane count and is honored FIRST;
    the automatic engine then optimizes WITHIN that structure (it only makes
    quota-preserving swaps and never moves must_include units). If the quotas force a
    barcode collision that no swap can resolve, it is reported and the file is not written.

Objective priority (high -> low):
  0. User --constraints (quotas + must_include)         (hard, if provided)
  1. No barcode conflicts within a lane                 (hard)
  2. Cycle 1/2 color feasibility (i7 & i5)              (--c12-weight, default 1e6)
  3. Balanced per-lane ratio sums (loading)            (--balance-weight, default 10)
  4. Cycle >=3 color balance                           (--w34 weighting)
  With realistic (color-balanced) index kits, step 2 is satisfied in any layout, so
  lanes come out ratio-balanced; with pathological indexes, color feasibility wins
  over balance. Set --balance-weight 0 for v2-style behavior (fixed round-robin lane
  sizes, color-only optimization).

Hard constraint definition (NEW emphasis):
  - Barcode uniqueness uses the FULL index strings (after optional --rc-i5), which is
    what actually matters for demultiplexing. (v2 only compared the truncated
    color-check window, which could over-report.)

Input CSV/TSV/XLSX must have headers: id,i7,i5  plus optional: ratio,type,lane,bundle
i5 is NOT reverse-complemented by default. Use --rc-i5 if needed.

Easy run
========
  # Assign into N lanes (auto-separates duplicate barcodes; honors `bundle`):
  python index_balance_v3.py group library.csv --lanes 2 --out lanes.csv
  # If 2 lanes is infeasible, it prints min-lanes-needed + removal plans.
  # Add --auto-expand to instead solve at the minimum feasible number of lanes.

  # Check an existing lane layout:
  python index_balance_v3.py check lanes.csv
  # headers: id,i7,i5,[ratio],[lane]

Outputs (human-readable):
  - check: per-lane color fractions + pass/fail (i7/i5), barcode-duplicate report,
           and the min pairwise Hamming distance within each lane.
  - group: feasibility verdict; if feasible, the lane summary + per-cycle detail and
           ratio balance; if infeasible, min-lanes-needed and removal plans.
"""

import argparse
import csv
import json
import random
from itertools import combinations, product
from math import prod
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from collections import Counter, defaultdict, OrderedDict, deque

try:
    import openpyxl
except Exception:
    openpyxl = None

BASES = ("A", "C", "G", "T")

# ============================================================================
# IO & utils
# ============================================================================

def revcomp(seq: str) -> str:
    comp = str.maketrans("ACGTNacgtn", "TGCANtgcan")
    return seq.translate(comp)[::-1]


# Optional columns we will carry through (with defaults) if present.
_OPT_DEFAULTS = {"ratio": "1", "type": "ALL", "lane": "ALL", "bundle": ""}


def read_table(path: str) -> List[Dict[str, str]]:
    """
    Read CSV/TSV/XLSX with required headers id,i7,i5 and optional ratio,type,lane,bundle.
    - CSV/TSV: delimiter by suffix (.csv => ',', otherwise '\\t'), encoding utf-8-sig (BOM-safe)
    - XLSX/XLSM: openpyxl, first sheet, first non-empty row as header
    Returns one dict per data row with keys: id,i7,i5,ratio(float),type,lane,bundle
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Input file not found: {path}")
    suffix = p.suffix.lower()
    items: List[Dict[str, str]] = []

    def make_record(getter) -> Optional[Dict[str, str]]:
        rec = {
            "id": getter("id"),
            "i7": getter("i7").upper(),
            "i5": getter("i5").upper(),
        }
        if not (rec["id"] or rec["i7"] or rec["i5"]):
            return None  # skip blank line
        try:
            rec["ratio"] = float(getter("ratio") or "1")
        except ValueError:
            raise ValueError(f"{rec['id']}: ratio must be a number, got '{getter('ratio')}'.")
        rec["type"] = getter("type") or "ALL"
        rec["lane"] = getter("lane") or ""
        rec["bundle"] = getter("bundle") or ""
        return rec

    # ---- Excel ----
    if suffix in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise ImportError("Reading XLSX/XLSM requires 'openpyxl'. Install: pip install openpyxl")
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        header = None
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            vals = [("" if v is None else str(v)).strip() for v in row]
            if any(v != "" for v in vals):
                header = vals
                break
        if not header:
            raise ValueError("No header row found in the Excel sheet.")
        header = [h.replace("﻿", "").strip() for h in header]
        colmap = {name.lower(): idx for idx, name in enumerate(header)}
        missing = {"id", "i7", "i5"} - set(colmap.keys())
        if missing:
            raise ValueError(f"Missing required columns: {missing}. Found: {header}")

        for row in ws.iter_rows(min_row=ws.min_row + 1, values_only=True):
            if row is None:
                continue
            def get(col):
                if col not in colmap:
                    return _OPT_DEFAULTS.get(col, "")
                idx = colmap[col]
                v = "" if idx >= len(row) or row[idx] is None else str(row[idx])
                return v.strip()
            rec = make_record(get)
            if rec is not None:
                items.append(rec)
        wb.close()
        return items

    # ---- CSV/TSV ----
    delim = "," if suffix == ".csv" else "\t"
    with p.open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=delim)
        fieldnames = [fn.replace("﻿", "").strip() for fn in (reader.fieldnames or [])]
        reader.fieldnames = fieldnames
        lower = {fn.lower(): fn for fn in fieldnames}
        missing = {"id", "i7", "i5"} - set(lower.keys())
        if missing:
            raise ValueError(f"Missing required columns: {missing}. Found: {fieldnames}")
        for row in reader:
            def get(col):
                if col in lower:
                    return (row.get(lower[col]) or "").strip()
                return _OPT_DEFAULTS.get(col, "")
            rec = make_record(get)
            if rec is not None:
                items.append(rec)
    return items


def validate_indexes(items: List[Dict[str, str]]) -> None:
    seen_ids = set()
    for r in items:
        if r["id"] in seen_ids:
            raise ValueError(f"Duplicate sample id '{r['id']}' in input (ids must be unique).")
        seen_ids.add(r["id"])
        for k in ("i7", "i5"):
            s = r[k]
            if not s:
                raise ValueError(f"{r['id']} has empty {k}.")
            for ch in s:
                if ch not in "ACGTN":
                    raise ValueError(f"{r['id']} {k} contains invalid base '{ch}'. Only A/C/G/T/N allowed.")


def effective_cycles(items: List[Dict[str, str]], which: str, requested: int) -> int:
    min_len = min(len(r[which]) for r in items) if items else 0
    return max(0, min(requested, min_len))


# ============================================================================
# Color math (unchanged from v2)
# ============================================================================

def per_cycle_base_lists(seqs: List[str], ratios: List[float], L: int) -> List[Dict[str, float]]:
    """Collect ratio-weighted base fractions per cycle (ignore N)."""
    cycles = [{'A': 0.0, 'C': 0.0, 'G': 0.0, 'T': 0.0} for _ in range(L)]
    for j in range(len(seqs)):
        s = seqs[j]
        r = ratios[j]
        for i in range(L):
            if i < len(s):
                b = s[i]
                if b in BASES:
                    cycles[i][b] += r
    for i in range(L):
        tot = sum(cycles[i].values())
        if tot == 0:
            continue
        for b in BASES:
            cycles[i][b] /= tot
    return cycles


def color_fractions_one_cycle(base_list: Dict[str, float]) -> Dict[str, float]:
    a, t, c, g = base_list['A'], base_list['T'], base_list['C'], base_list['G']
    tot_units = 2 * (a + t + c + g)
    if tot_units == 0:
        return {"green": 0.0, "blue": 0.0, "dark": 0.0}
    green = (t * 2.0 + c * 1.0) / tot_units
    blue = (a * 2.0 + c * 1.0) / tot_units
    dark = (g * 2.0) / tot_units
    return {"green": green, "blue": blue, "dark": dark}


def hard_cycles_ok_aggregate(seqs, ratios, L, min_green, min_blue, max_dark):
    blists = per_cycle_base_lists(seqs, ratios, L)
    colors_12: List[Optional[Dict[str, float]]] = [None, None]
    for idx in (0, 1):
        if L > idx:
            col = color_fractions_one_cycle(blists[idx])
            colors_12[idx] = col
            ok = (col["green"] >= min_green) and (col["blue"] > min_blue) and (col["dark"] <= max_dark)
            if not ok:
                return False, colors_12
    return True, colors_12


def make_weights_special(L: int, w34: float) -> List[float]:
    if L <= 2:
        return [0.0] * L
    w34 = min(1.0, max(0.0, w34))
    w = [0.0] * L
    if 3 <= L <= 4:
        n = L - 2
        for i in range(2, L):
            w[i] = 1.0 / n
    else:
        w[2] = w34 / 2.0
        w[3] = w34 / 2.0
        n_tail = L - 4
        if n_tail > 0:
            tail_share = (1.0 - w34) / n_tail
            for i in range(4, L):
                w[i] = tail_share
    total = sum(w)
    if total > 0:
        w = [x / total for x in w]
    return w


def weighted_penalty(seqs, ratios, L, min_green, min_blue, max_dark, weights) -> float:
    if L <= 0 or not seqs or not weights:
        return 0.0
    blists = per_cycle_base_lists(seqs, ratios, L)
    pen = 0.0
    for i in range(min(L, len(weights))):
        if weights[i] == 0.0:
            continue
        col = color_fractions_one_cycle(blists[i])
        v = 0.0
        v += max(0.0, (min_green - col["green"]))
        v += max(0.0, (min_blue - col["blue"]))
        v += max(0.0, (col["dark"] - max_dark))
        pen += weights[i] * v
    return pen


def c12_violation_aggregate(seqs, ratios, L, min_green, min_blue, max_dark, strict_blue=True) -> float:
    if not seqs or L <= 0:
        return 0.0
    EPS = 1e-9 if strict_blue else 0.0
    blists = per_cycle_base_lists(seqs, ratios, L)
    pen = 0.0
    for idx in (0, 1):
        if L > idx:
            col = color_fractions_one_cycle(blists[idx])
            v = 0.0
            v += max(0.0, (min_green - col["green"]))
            v += max(0.0, (min_blue - col["blue"]) + EPS)
            v += max(0.0, (col["dark"] - max_dark))
            pen += v
    return pen


def evaluate_color_balance_strict_weighted(seqs, ratios, cycles, min_green, min_blue, max_dark, w34) -> Dict[str, Any]:
    hard_ok, colors_12 = hard_cycles_ok_aggregate(seqs, ratios, cycles, min_green, min_blue, max_dark)
    blists = per_cycle_base_lists(seqs, ratios, cycles)
    per_cycle: List[Dict[str, Any]] = []
    worst = None
    worst_score = -1.0
    for i, bl in enumerate(blists, start=1):
        col = color_fractions_one_cycle(bl)
        ok = (col["green"] >= min_green) and (col["blue"] > min_blue) and (col["dark"] <= max_dark)
        per_cycle.append({"cycle": i, "color": col, "ok": ok})
        v = max(0.0, (min_green - col["green"])) + max(0.0, (min_blue - col["blue"])) + max(0.0, (col["dark"] - max_dark))
        if v > worst_score:
            worst_score = v
            worst = {"cycle": i, "color": col, "color_ok": (v == 0.0)}
    denom = cycles if cycles > 0 else 1
    mean_color = {
        "green": (sum(pc["color"]["green"] for pc in per_cycle) / denom) if per_cycle else 0.0,
        "blue": (sum(pc["color"]["blue"] for pc in per_cycle) / denom) if per_cycle else 0.0,
        "dark": (sum(pc["color"]["dark"] for pc in per_cycle) / denom) if per_cycle else 0.0,
    }
    weights = make_weights_special(cycles, w34)
    wpen = weighted_penalty(seqs, ratios, cycles, min_green, min_blue, max_dark, weights)
    return {
        "cycles": cycles, "hard_ok": hard_ok, "hard_c12": colors_12,
        "per_cycle": per_cycle, "weighted_penalty": wpen, "weights": weights,
        "mean_color": mean_color, "worst_cycle": worst,
    }


def min_pairwise_hamming(seqs: List[str]) -> Optional[int]:
    """Min pairwise Hamming distance over equal-length sequences (None if <2 or ragged)."""
    if len(seqs) < 2:
        return None
    if len({len(s) for s in seqs}) != 1:
        return None
    best = None
    for i in range(len(seqs)):
        for j in range(i + 1, len(seqs)):
            d = sum(1 for a, b in zip(seqs[i], seqs[j]) if a != b)
            if best is None or d < best:
                best = d
                if best == 0:
                    return 0
    return best


# ============================================================================
# Units / bundles
# ============================================================================

def build_units(items: List[Dict[str, str]], rc_i5: bool):
    """
    Group rows into atomic UNITS by `bundle`. Rows with empty bundle are singletons.
    Each unit carries full (i7, rc?i5) barcode pairs for conflict detection.
    Returns (units, internal_conflicts).
      unit = {
        uid, is_bundle, members(list of records),
        full_pairs(list of (i7_full, i5_full_after_rc)),
        ratio(sum), nsamp, type(common type or None if mixed)
      }
      internal_conflicts = list of (uid, members, dup_pairs) for bundles that collide internally.
    """
    groups: "OrderedDict[str, List[Dict[str,str]]]" = OrderedDict()
    for r in items:
        b = (r.get("bundle") or "").strip()
        key = b if b else f"__solo__::{r['id']}"
        groups.setdefault(key, []).append(r)

    units = []
    internal_conflicts = []
    for key, members in groups.items():
        is_bundle = bool((members[0].get("bundle") or "").strip())
        uid = (members[0]["bundle"].strip() if is_bundle else members[0]["id"])
        # Ballast classification is UNIT-level: a unit is ballast (ratio-only, no barcode -> no
        # conflict/color) iff ALL its members lack a full i7+i5. A bundle that MIXES indexed and
        # no-index members is rejected, so indexed members are never silently dropped.
        n_full = sum(bool(m["i7"]) and bool(m["i5"]) for m in members)
        if n_full == len(members):
            full_pairs = [(m["i7"], revcomp(m["i5"]) if rc_i5 else m["i5"]) for m in members]
            if len(set(full_pairs)) != len(full_pairs):
                cnt = Counter(full_pairs)
                dups = [p for p, n in cnt.items() if n > 1]
                internal_conflicts.append((uid, members, dups))
            is_ballast = False
        elif n_full == 0:
            full_pairs = []
            is_ballast = True
        else:
            raise ValueError(f"bundle '{uid}' mixes indexed and no-index members; a bundle must "
                             f"be either all-indexed or all no-index (ballast).")
        typecounts = Counter(m.get("type", "ALL") for m in members)
        units.append({
            "uid": uid,
            "is_bundle": is_bundle,
            "is_ballast": is_ballast,                          # no-index ratio-only pool
            "members": members,
            "full_pairs": full_pairs,
            "ratio": sum(m["ratio"] for m in members),
            "nsamp": len(members),
            "type": (next(iter(typecounts)) if len(typecounts) == 1 else None),
            "typecounts": typecounts,                          # type -> sample count (for quotas)
            "typesig": tuple(sorted(typecounts.items())),      # identity for quota-preserving swaps
        })
    return units, internal_conflicts


def finalize_units(units, L7, L5, rc_i5):
    """Attach truncated color-check sequences used for color-balance scoring."""
    for u in units:
        if u.get("is_ballast"):
            u["m_i7"], u["m_i5"], u["m_ratio"] = [], [], []   # no color contribution
            continue
        m_i7, m_i5, m_ratio = [], [], []
        for m in u["members"]:
            i7 = m["i7"][:L7]
            i5raw = revcomp(m["i5"]) if rc_i5 else m["i5"]
            m_i7.append(i7)
            m_i5.append(i5raw[:L5])
            m_ratio.append(m["ratio"])
        u["m_i7"], u["m_i5"], u["m_ratio"] = m_i7, m_i5, m_ratio


# ============================================================================
# Conflict graph + coloring
# ============================================================================

def build_index_map(units) -> "defaultdict[Tuple[str,str], List[int]]":
    """barcode pair -> list of unit indices containing it."""
    idx_to_units = defaultdict(list)
    for i, u in enumerate(units):
        for p in set(u["full_pairs"]):
            idx_to_units[p].append(i)
    return idx_to_units


def build_adjacency(n: int, idx_to_units) -> List[set]:
    adj = [set() for _ in range(n)]
    for _, us in idx_to_units.items():
        if len(us) > 1:
            for a in range(len(us)):
                for b in range(a + 1, len(us)):
                    adj[us[a]].add(us[b])
                    adj[us[b]].add(us[a])
    return adj


def max_multiplicity(idx_to_units) -> int:
    return max((len(us) for us in idx_to_units.values()), default=1)


class _BudgetExceeded(Exception):
    pass


def _exact_color(adj: List[set], k: int, budget: int) -> Tuple[Optional[List[int]], bool]:
    """Backtracking k-coloring with symmetry breaking + step budget.
    Returns (coloring or None, decisive). decisive=False means budget ran out (unknown)."""
    n = len(adj)
    if n == 0:
        return [], True
    order = sorted(range(n), key=lambda i: len(adj[i]), reverse=True)
    color = [-1] * n
    steps = [0]

    def bt(pos: int, max_used: int) -> bool:
        if pos == n:
            return True
        steps[0] += 1
        if steps[0] > budget:
            raise _BudgetExceeded()
        v = order[pos]
        forbidden = {color[u] for u in adj[v] if color[u] >= 0}
        upto = min(k - 1, max_used + 1)
        for c in range(0, upto + 1):
            if c in forbidden:
                continue
            color[v] = c
            if bt(pos + 1, max(max_used, c)):
                return True
            color[v] = -1
        return False

    try:
        ok = bt(0, -1)
        return (color[:] if ok else None), True
    except _BudgetExceeded:
        return None, False


def _dsatur_color(adj: List[set], k: int) -> Optional[List[int]]:
    """Greedy DSATUR; returns a coloring using <= k colors, or None if it needs more."""
    n = len(adj)
    color = [-1] * n
    sat = [set() for _ in range(n)]
    deg = [len(adj[i]) for i in range(n)]
    for _ in range(n):
        best, best_key = -1, None
        for v in range(n):
            if color[v] >= 0:
                continue
            key = (len(sat[v]), deg[v])
            if best_key is None or key > best_key:
                best_key, best = key, v
        v = best
        used = {color[u] for u in adj[v] if color[u] >= 0}
        c = 0
        while c in used:
            c += 1
        if c >= k:
            return None
        color[v] = c
        for u in adj[v]:
            sat[u].add(c)
    return color


def color_units(adj: List[set], k: int, budget: int = 400000) -> Tuple[Optional[List[int]], bool]:
    """Try to color the conflict graph with k colors.
    Returns (coloring or None, exact). exact=True means the answer is certain."""
    coloring, decisive = _exact_color(adj, k, budget)
    if decisive:
        return coloring, True
    # budget exceeded -> best-effort heuristic
    ds = _dsatur_color(adj, k)
    return ds, False


def min_lanes_needed(adj: List[set], idx_to_units, cap: Optional[int] = None) -> Tuple[int, bool]:
    """Smallest k for which the graph is k-colorable. Returns (k, exact)."""
    n = len(adj)
    lb = max_multiplicity(idx_to_units)
    hi = cap if cap is not None else max(lb, n)
    overall_exact = True
    for k in range(lb, hi + 1):
        coloring, exact = color_units(adj, k)
        if not exact:
            overall_exact = False
        if coloring is not None:
            return k, (exact and overall_exact)
    return hi, overall_exact


# ============================================================================
# Assignment: initial layout + conflict repair + color/ratio optimization
# ============================================================================

def _lane_pairs(bucket) -> List[Tuple[str, str]]:
    return [p for u in bucket for p in u["full_pairs"]]


def lane_has_conflict(bucket) -> bool:
    pairs = _lane_pairs(bucket)
    return len(set(pairs)) != len(pairs)


def any_conflict(buckets) -> bool:
    return any(lane_has_conflict(b) for b in buckets)


def initial_round_robin_units(units, L: int) -> List[list]:
    """Round-robin units sorted by barcode key so duplicates land in different lanes."""
    order = sorted(range(len(units)), key=lambda i: (min(units[i]["full_pairs"]), str(units[i]["uid"])))
    buckets = [[] for _ in range(L)]
    for k, i in enumerate(order):
        buckets[k % L].append(units[i])
    return buckets


def repair_conflicts(buckets, L: int, max_passes: int = 1000) -> bool:
    """Relocate/swap conflicting units until every lane is conflict-free (best effort)."""
    for _ in range(max_passes):
        conflict_exists = False
        moved = False
        for li in range(L):
            lane = buckets[li]
            cnt = Counter(_lane_pairs(lane))
            dup = {p for p, n in cnt.items() if n > 1}
            if not dup:
                continue
            conflict_exists = True
            cu_idx = next((k for k, u in enumerate(lane) if any(p in dup for p in u["full_pairs"])), None)
            if cu_idx is None:
                continue
            cu = lane[cu_idx]
            # try a plain move to a lane sharing no barcode with cu
            dest = None
            for lj in sorted(range(L), key=lambda x: sum(uu["ratio"] for uu in buckets[x])):
                if lj == li:
                    continue
                other = {p for uu in buckets[lj] for p in uu["full_pairs"]}
                if not any(p in other for p in cu["full_pairs"]):
                    dest = lj
                    break
            if dest is not None:
                lane.pop(cu_idx)
                buckets[dest].append(cu)
                moved = True
                continue
            # else try a swap that leaves both lanes conflict-free
            li_pairs = {p for x, u in enumerate(lane) if x != cu_idx for p in u["full_pairs"]}
            done = False
            for lj in range(L):
                if lj == li:
                    continue
                for kk, vv in enumerate(buckets[lj]):
                    lj_pairs = {p for x, u in enumerate(buckets[lj]) if x != kk for p in u["full_pairs"]}
                    if (not any(p in li_pairs for p in vv["full_pairs"])) and \
                       (not any(p in lj_pairs for p in cu["full_pairs"])):
                        lane[cu_idx], buckets[lj][kk] = vv, cu
                        moved = True
                        done = True
                        break
                if done:
                    break
        if not conflict_exists:
            return True
        if not moved:
            break
    return not any_conflict(buckets)


def lane_cost(bucket, L7, L5, mg, mb, md, w7, w5, c12_weight, target_ratio, balance_weight) -> float:
    """Per-lane cost: conflict (hard) -> C1/C2 violations -> weighted >=3 -> ratio balance (tertiary).

    Ballast units (no-index pools) count toward the lane's read budget (ratio balance) but
    are excluded from conflict and color. A lane that contains any ballast member is treated
    as 'complex pool covered' and its color is NOT scored (color is optimized only for lanes
    whose members are all fully indexed)."""
    i7s, i5s, rr, pairs = [], [], [], []
    ratio_sum = 0.0
    has_ballast = False
    for u in bucket:
        ratio_sum += u["ratio"]                  # full read budget incl. ballast
        if u.get("is_ballast"):
            has_ballast = True
            continue                             # no barcode, no color
        i7s += u["m_i7"]; i5s += u["m_i5"]; rr += u["m_ratio"]; pairs += u["full_pairs"]
    if len(set(pairs)) != len(pairs):
        return 1e12  # barcode collision in this lane: forbidden
    if has_ballast:                              # pool-covered lane: skip color scoring
        c12 = 0.0; pen = 0.0
    else:
        c12 = (c12_violation_aggregate(i7s, rr, L7, mg, mb, md, True) +
               c12_violation_aggregate(i5s, rr, L5, mg, mb, md, True))
        pen = (weighted_penalty(i7s, rr, L7, mg, mb, md, w7) +
               weighted_penalty(i5s, rr, L5, mg, mb, md, w5))
    bal = 0.0
    if balance_weight > 0 and target_ratio > 0:
        bal = balance_weight * abs(ratio_sum - target_ratio) / target_ratio
    return c12_weight * c12 + pen + bal


def optimize_units(buckets, L7, L5, rc_i5, iters, seed, mg, mb, md, w7, w5,
                   c12_weight=1e6, balance_weight=0.05,
                   locked_uids=None, allow_moves=True,
                   restrict_same_type=False, restrict_same_size=False,
                   quota_lock=False) -> List[list]:
    """Randomized swap/move optimizer over whole units; conflicts are rejected via the cost.

    quota_lock=True: only swap units with identical type composition (and never move), so
    per-lane per-type counts — i.e. user --constraints quotas — are preserved exactly."""
    random.seed(seed)
    L = len(buckets)
    locked_uids = locked_uids or set()
    if quota_lock:
        allow_moves = False  # moves change per-lane counts and would break quotas
    total_ratio = sum(u["ratio"] for b in buckets for u in b)
    target = total_ratio / L if L else 0.0

    def cost(b):
        return lane_cost(b, L7, L5, mg, mb, md, w7, w5, c12_weight, target, balance_weight)

    lane_scores = [cost(b) for b in buckets]
    total = sum(lane_scores)

    def movable(u):
        return u["uid"] not in locked_uids

    for _ in range(iters):
        do_move = allow_moves and (random.random() < 0.5)

        if do_move:
            a = random.randrange(L)
            if len(buckets[a]) <= 1:
                continue  # never empty a lane: every lane stays used
            ia = random.randrange(len(buckets[a]))
            if not movable(buckets[a][ia]):
                continue
            b = random.randrange(L)
            if b == a:
                continue
            u = buckets[a].pop(ia)
            buckets[b].append(u)
            new_a, new_b = cost(buckets[a]), cost(buckets[b])
            new_total = total - lane_scores[a] - lane_scores[b] + new_a + new_b
            if new_total < total - 1e-12:
                lane_scores[a], lane_scores[b] = new_a, new_b
                total = new_total
            else:
                buckets[b].pop()
                buckets[a].insert(ia, u)
            continue

        # swap
        a, b = random.randrange(L), random.randrange(L)
        if a == b or not buckets[a] or not buckets[b]:
            continue
        ia, ib = random.randrange(len(buckets[a])), random.randrange(len(buckets[b]))
        ua, ub = buckets[a][ia], buckets[b][ib]
        if not movable(ua) or not movable(ub):
            continue
        if restrict_same_type and (ua["type"] != ub["type"]):
            continue
        if restrict_same_size and (ua["nsamp"] != ub["nsamp"]):
            continue
        if quota_lock and (ua["typesig"] != ub["typesig"]):
            continue  # keep each lane's per-type counts (quotas) unchanged
        buckets[a][ia], buckets[b][ib] = ub, ua
        new_a, new_b = cost(buckets[a]), cost(buckets[b])
        new_total = total - lane_scores[a] - lane_scores[b] + new_a + new_b
        if new_total < total - 1e-12:
            lane_scores[a], lane_scores[b] = new_a, new_b
            total = new_total
        else:
            buckets[a][ia], buckets[b][ib] = ua, ub

    return buckets


def fill_empty_lanes(buckets, L) -> None:
    """One-time: move units from the largest lanes into empty lanes (always conflict-free)
    so every lane is used when there are at least L units."""
    for _ in range(L * 4):
        empties = [i for i in range(L) if not buckets[i]]
        if not empties:
            return
        donors = sorted((i for i in range(L) if len(buckets[i]) > 1),
                        key=lambda i: -len(buckets[i]))
        if not donors:
            return  # fewer units than lanes; cannot fill all
        moved = False
        for e in empties:
            for d in donors:
                if len(buckets[d]) > 1:
                    buckets[e].append(buckets[d].pop())  # moving into an empty lane never conflicts
                    moved = True
                    break
        if not moved:
            return


def assign_and_optimize(units, L, L7, L5, rc_i5, iters, seed, mg, mb, md, w7, w5,
                        c12_weight, balance_weight) -> Optional[List[list]]:
    """Build a conflict-free, lane-filled layout for `units` into L lanes, then optimize.

    Lane sizes are seeded by a balanced round-robin (duplicates land in different lanes when
    multiplicity <= L). Optimization then runs swaps plus, when balance_weight > 0, unit
    moves (guarded so a lane is never emptied) to balance per-lane ratio sums; with
    balance_weight == 0 it is swap-only (fixed round-robin sizes, color-only). Returns None
    if no conflict-free layout exists for L lanes.
    """
    idx_to_units = build_index_map(units)
    adj = build_adjacency(len(units), idx_to_units)

    buckets = initial_round_robin_units(units, L)
    if not repair_conflicts(buckets, L):
        coloring, _ = color_units(adj, L)
        if coloring is None:
            return None
        buckets = [[] for _ in range(L)]
        for i, c in enumerate(coloring):
            buckets[c].append(units[i])

    fill_empty_lanes(buckets, L)
    # balance_weight > 0: enable unit moves so lane ratio sums can be balanced
    #   (priority: conflict >> cycle1/2 color >> ratio balance >> cycle>=3 color).
    # balance_weight == 0: swap-only, keeping the round-robin's fixed balanced lane sizes
    #   and optimizing color alone (v2-style).
    buckets = optimize_units(buckets, L7, L5, rc_i5, iters, seed, mg, mb, md, w7, w5,
                             c12_weight=c12_weight, balance_weight=balance_weight,
                             allow_moves=(balance_weight > 0))
    return buckets


# ============================================================================
# Manual constraints (per-lane quotas & must_include) — highest priority
# ============================================================================

def load_constraints(path: str) -> dict:
    """Load a JSON or YAML constraints file with a top-level 'lanes' mapping."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Constraints file not found: {path}")
    suf = p.suffix.lower()
    if suf == ".json":
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
    elif suf in (".yml", ".yaml"):
        try:
            import yaml
        except Exception:
            raise RuntimeError("Reading YAML constraints requires PyYAML. Use JSON, or: pip install pyyaml")
        with open(p, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
    else:
        raise ValueError("Constraints must be .json or .yaml/.yml")
    if not isinstance(data, dict) or "lanes" not in data or not isinstance(data["lanes"], dict):
        raise ValueError("Constraints must have a top-level 'lanes' mapping.")
    return data


def normalize_constraints(data: dict):
    """Return (lane_ids[sorted int], lane_quotas{lid: {type:count}}, lane_must{lid: set(ids)})."""
    lanes_cfg = data["lanes"]
    lane_ids = sorted(int(k) for k in lanes_cfg.keys())
    lane_quotas, lane_must = {}, {}
    for lid in lane_ids:
        cfg = lanes_cfg.get(str(lid), lanes_cfg.get(lid)) or {}
        quotas = cfg.get("quotas", {}) or {}
        must = cfg.get("must_include", []) or []
        if not isinstance(quotas, dict):
            raise ValueError(f"lane {lid}: 'quotas' must be a dict of type->count.")
        if not all(isinstance(v, int) and v >= 0 for v in quotas.values()):
            raise ValueError(f"lane {lid}: quota counts must be non-negative integers.")
        lane_quotas[lid] = {str(t): int(c) for t, c in quotas.items()}
        lane_must[lid] = {str(s) for s in must}
    return lane_ids, lane_quotas, lane_must


def check_quota_feasibility(items, lane_ids, lane_quotas) -> None:
    """Sample-level sanity: total quotas == #samples, and enough samples of each type."""
    avail = Counter(r.get("type", "ALL") for r in items)
    need = Counter()
    for lid in lane_ids:
        for t, c in lane_quotas[lid].items():
            need[t] += c
    if sum(need.values()) != len(items):
        raise ValueError(f"Sum of all lane quotas ({sum(need.values())}) must equal the number "
                         f"of input samples ({len(items)}).")
    for t, n in need.items():
        if avail.get(t, 0) < n:
            raise ValueError(f"Not enough samples of type '{t}': need {n}, available {avail.get(t, 0)}.")


def place_units_with_quotas(units, lane_ids, lane_quotas, lane_must, id_to_unit) -> List[list]:
    """Place atomic units into lanes so each lane's per-type sample counts EXACTLY meet its
    quota, honoring must_include. Prefers conflict-free lanes. Bundles stay intact (a bundle
    contributes its members' types to whatever lane it lands in). Raises on infeasible quotas."""
    L = len(lane_ids)
    buckets = [[] for _ in lane_ids]
    remaining = [Counter(lane_quotas[lid]) for lid in lane_ids]
    placed_lane = {}   # unit index -> lane id it was pinned to (must_include)

    def fits(rem, tc):
        return all(rem.get(t, 0) >= n for t, n in tc.items())

    # must_include first (id may be a sample id or a bundle id; multiple member ids of one
    # bundle resolve to the same unit — pinning it repeatedly to the SAME lane is harmless).
    for i, lid in enumerate(lane_ids):
        for sid in sorted(lane_must[lid]):
            if sid not in id_to_unit:
                raise ValueError(f"lane {lid} must_include '{sid}' not found (sample id or bundle id).")
            k = id_to_unit[sid]
            if k in placed_lane:
                if placed_lane[k] == lid:
                    continue   # redundant pin of the same unit to the same lane: ignore
                raise ValueError(f"must_include '{sid}' pins bundle '{units[k]['uid']}' to lane {lid}, "
                                 f"but it is already pinned to lane {placed_lane[k]} (a bundle cannot span lanes).")
            tc = units[k]["typecounts"]
            if not fits(remaining[i], tc):
                raise ValueError(f"lane {lid} quota cannot hold must_include '{sid}' "
                                 f"(types {dict(tc)}).")
            buckets[i].append(units[k]); remaining[i].subtract(tc); placed_lane[k] = lid

    # fill the rest by backtracking (largest units first), preferring conflict-free lanes
    rest = sorted((k for k in range(len(units)) if k not in placed_lane),
                  key=lambda k: (-units[k]["nsamp"], units[k]["uid"]))

    def conflicts(i, u):
        seen = {p for uu in buckets[i] for p in uu["full_pairs"]}
        return any(p in seen for p in u["full_pairs"])

    def bt(j):
        if j == len(rest):
            return all(sum(r.values()) == 0 for r in remaining)
        u = units[rest[j]]
        tc = u["typecounts"]
        order = sorted(range(L), key=lambda i: (conflicts(i, u), -sum(remaining[i].values())))
        for i in order:
            if fits(remaining[i], tc):
                buckets[i].append(u); remaining[i].subtract(tc)
                if bt(j + 1):
                    return True
                buckets[i].pop(); remaining[i].update(tc)
        return False

    if not bt(0):
        raise ValueError("Cannot satisfy lane quotas with the given samples/bundles "
                         "(quota counts incompatible with bundle sizes/types).")
    return buckets


# ============================================================================
# Removal-plan generation (when infeasible at the requested #lanes)
# ============================================================================

def _ensure_colorable(units, removed, L) -> set:
    """Augment `removed` (set of unit idx) until the rest is L-colorable. Returns final set."""
    removed = set(removed)
    while True:
        kept_idx = [i for i in range(len(units)) if i not in removed]
        sub = [units[i] for i in kept_idx]
        sub_adj = build_adjacency(len(sub), build_index_map(sub))
        coloring, _ = color_units(sub_adj, L)
        if coloring is not None:
            return removed
        # drop the highest-degree remaining unit (prefer low ratio on ties)
        j = min(range(len(sub)), key=lambda x: (-len(sub_adj[x]), units[kept_idx[x]]["ratio"]))
        removed.add(kept_idx[j])


def generate_removal_plans(units, L: int, max_plans: int = 3):
    """
    Propose minimal sets of units to REMOVE so the rest fit into L lanes conflict-free.
    Returns a list of plans, each a sorted list of removed unit indices (into `units`),
    ranked by least material (ratio) dropped. Distinct choices within a barcode group
    are surfaced as separate plans (e.g. drop sample A vs B vs C).
    """
    idx_to_units = build_index_map(units)
    adj = build_adjacency(len(units), idx_to_units)
    over = {p: list(us) for p, us in idx_to_units.items() if len(us) > L}
    unit_over = defaultdict(set)
    for p, us in over.items():
        for u in us:
            unit_over[u].add(p)
    coupled = any(len(v) > 1 for v in unit_over.values())  # a unit on >1 over-subscribed barcode

    plans, seen = [], set()

    def add_plan(rem_set):
        rem = _ensure_colorable(units, rem_set, L)  # always returns an L-colorable removal set
        key = frozenset(rem)
        if rem and key not in seen:
            seen.add(key)
            plans.append(sorted(rem))

    # (a) Over-subscribed barcodes (multiplicity > L): enumerate minimal, distinct drop
    # combinations ranked by material lost, so users see real alternatives (drop A vs B vs C).
    if over:
        choices = []
        for p, us in over.items():
            need = len(us) - L
            combos = sorted(combinations(us, need), key=lambda c: sum(units[u]["ratio"] for u in c))
            choices.append(combos)
        enum_size = prod(len(c) for c in choices) if choices else 0

        if not coupled and 0 < enum_size <= 5000:
            candidates = sorted(
                {frozenset(u for grp in combo for u in grp) for combo in product(*choices)},
                key=lambda rem: (sum(units[u]["ratio"] for u in rem), len(rem)))
            for rem in candidates:
                add_plan(rem)
                if len(plans) >= max_plans:
                    break
        else:
            # Coupled / very large: greedy multi-hitting-set with a few tie-break strategies.
            def cover(u):
                return len(unit_over[u])

            def greedy(keyfn) -> set:
                counts = {p: len(us) for p, us in over.items()}
                active = set(unit_over.keys())
                removed = set()
                while any(counts[p] > L for p in over):
                    cands = [u for u in active if any(counts[p] > L for p in unit_over[u])]
                    if not cands:
                        break
                    u = min(cands, key=keyfn)
                    active.discard(u)
                    removed.add(u)
                    for p in unit_over[u]:
                        counts[p] -= 1
                return removed

            for keyfn in (
                lambda u: (units[u]["ratio"], -cover(u), units[u]["nsamp"]),   # preserve material
                lambda u: (-cover(u), units[u]["ratio"], units[u]["nsamp"]),   # fewest removals
                lambda u: (units[u]["nsamp"], units[u]["ratio"], -cover(u)),   # fewest lost samples
            ):
                add_plan(greedy(keyfn))
                if len(plans) >= max_plans:
                    break

    # (b) Structural infeasibility + diversity top-up. When the conflict graph needs > L
    # colors purely from its structure (e.g. an odd cycle of bundles where every barcode
    # multiplicity is <= L, so step (a) finds nothing), seed a removal from each
    # conflict-participating unit and let _ensure_colorable() complete it. This is what
    # produces "drop B1 / B2 / B3" for a triangle that step (a) cannot see.
    if len(plans) < max_plans:
        seeds = sorted((i for i in range(len(units)) if adj[i]),
                       key=lambda i: (units[i]["ratio"], units[i]["nsamp"]))
        for u in seeds:
            add_plan({u})
            if len(plans) >= max_plans:
                break
        if not plans:                      # last resort: pure greedy removal from scratch
            add_plan(set())

    plans.sort(key=lambda rem: (sum(units[u]["ratio"] for u in rem), len(rem)))
    return plans[:max_plans]


# ============================================================================
# Pretty printing
# ============================================================================

def color_text(text: str, color: str = "yellow", bold: bool = False) -> str:
    codes = {"red": "31", "green": "32", "yellow": "33", "blue": "34", "magenta": "35", "cyan": "36"}
    return f"\033[{1 if bold else 0};{codes.get(color, '0')}m{text}\033[0m"


def fmt_bool(b: bool) -> str:
    return "PASS" if b else "FAIL"


def fmt_color_triplet(d: Dict[str, float]) -> str:
    return f"green={d['green']:.2f}  blue={d['blue']:.2f}  dark={d['dark']:.2f}"


def suggest_base_adjustment(color, min_green, min_blue, max_dark) -> str:
    sug = []
    if color.get("green", 0.0) < min_green:
        sug.append("↑(T/C) to raise green")
    if color.get("blue", 0.0) <= min_blue:
        sug.append("↑(A/C) to raise blue")
    if color.get("dark", 0.0) > max_dark:
        sug.append("↓G to reduce dark")
    return "" if not sug else " -> To PASS: " + "; ".join(sug)


def print_per_cycle_list(per_cycle, indent="  ", min_green=None, min_blue=None, max_dark=None) -> None:
    print(f"{indent}Per-cycle results:")
    for pc in per_cycle:
        status = "PASS" if pc["ok"] else "FAIL"
        extra = ""
        if (not pc["ok"]) and (min_green is not None):
            extra = suggest_base_adjustment(pc["color"], min_green, min_blue, max_dark)
        print(f"{indent}  #{pc['cycle']:>2}  {fmt_color_triplet(pc['color'])}  [{status}]{extra}")


def print_group_human(report, min_green, min_blue, max_dark, rc_i5, out_path="") -> None:
    print("=== Lane Assignment (Color Balance) ===")
    total_libs = sum(l["n_libraries"] for l in report["lane_reports"])
    print(f"Total libraries: {total_libs}")
    print(f"Cycles used: i7={report['cycles']['i7']}, i5={report['cycles']['i5']}")
    print(f"Number of lanes: {report['lanes']}\n")

    for lane in report["lane_reports"]:
        status = "PASS" if (lane["i7"]["hard_ok"] and lane["i5"]["hard_ok"]) else "FAIL"
        print(f"Lane {lane['lane']}: {status}  ({lane['n_libraries']} libraries, ratio sum={lane.get('ratio_sum', 0):.2f})")
        if lane.get("types"):
            print(f"  types: {dict(lane['types'])}")
        if lane.get("min_hamming") is not None:
            print(f"  min pairwise Hamming (i7+i5): {lane['min_hamming']}")
        print(f"  i7 mean: {fmt_color_triplet(lane['i7']['mean_color'])}")
        print(f"  i5 mean: {fmt_color_triplet(lane['i5']['mean_color'])}")
        print(f"  i7 weighted penalty (>=3): {lane['i7']['weighted_penalty']:.4f}")
        print(f"  i5 weighted penalty (>=3): {lane['i5']['weighted_penalty']:.4f}")
        print("  i7:")
        print_per_cycle_list(lane["i7"]["per_cycle"], "    ", min_green, min_blue, max_dark)
        print("  i5:")
        print_per_cycle_list(lane["i5"]["per_cycle"], "    ", min_green, min_blue, max_dark)
        print()

    if rc_i5:
        print(color_text("  Note: i5 was evaluated as reverse-complement (--rc-i5).", "cyan", True))
    if "ratio_balance" in report:
        rb = report["ratio_balance"]
        print(color_text(f"Ratio balance across lanes: min={rb['min']:.2f}  max={rb['max']:.2f}  "
                         f"spread={rb['spread']:.2f}  CV={rb['cv']:.3f}", "cyan"))
    print(color_text(f"All lanes pass (hard C1-2): {'YES' if report['all_lanes_hard_pass'] else 'NO'}",
                     "magenta", True))
    if out_path:
        print(f"Assignment saved to: {out_path}")


def ratio_balance_stats(buckets) -> Dict[str, float]:
    sums = [sum(u["ratio"] for u in b) for b in buckets]
    if not sums:
        return {"min": 0, "max": 0, "spread": 0, "cv": 0, "sums": []}
    mn, mx = min(sums), max(sums)
    mean = sum(sums) / len(sums)
    var = sum((s - mean) ** 2 for s in sums) / len(sums)
    cv = (var ** 0.5 / mean) if mean > 0 else 0.0
    return {"min": mn, "max": mx, "spread": mx - mn, "cv": cv, "sums": sums}


def build_lane_reports(buckets, L7, L5, rc_i5, mg, mb, md, w34):
    """Expand units to members and build per-lane color-balance reports."""
    lane_reports = []
    all_hard = True
    for lane_id, buck in enumerate(buckets, start=1):
        i7, i5, ratios = [], [], []
        combined = []
        type_counts = Counter()
        for u in buck:
            i7 += u["m_i7"]
            i5 += u["m_i5"]
            ratios += u["m_ratio"]
            for a, b in zip(u["m_i7"], u["m_i5"]):
                combined.append(a + b)
            for m in u["members"]:
                type_counts[m.get("type", "ALL")] += 1
        rep7 = evaluate_color_balance_strict_weighted(i7, ratios, L7, mg, mb, md, w34)
        rep5 = evaluate_color_balance_strict_weighted(i5, ratios, L5, mg, mb, md, w34)
        lane_reports.append({
            "lane": lane_id,
            "n_libraries": len(i7),
            "ratio_sum": sum(ratios),
            "i7": rep7, "i5": rep5,
            "types": dict(type_counts),
            "min_hamming": min_pairwise_hamming(combined),
        })
        all_hard = all_hard and rep7["hard_ok"] and rep5["hard_ok"]
    return lane_reports, all_hard


def print_plan(plan_no, removed_units, buckets, L7, L5, rc_i5, mg, mb, md, w34):
    n_samp = sum(u["nsamp"] for u in removed_units)
    r_sum = sum(u["ratio"] for u in removed_units)
    print(color_text(f"--- Removal Plan {plan_no}: drop {len(removed_units)} unit(s) "
                     f"/ {n_samp} sample(s), ratio={r_sum:.2f} ---", "yellow", True))
    for u in removed_units:
        tag = "BUNDLE" if u["is_bundle"] else "sample"
        ids = ",".join(m["id"] for m in u["members"])
        print(f"   remove {tag} {u['uid']}  (n={u['nsamp']}, ratio={u['ratio']:.2f})  ids: {ids}")
    rb = ratio_balance_stats(buckets)
    print(f"   Resulting {len(buckets)} lanes  (ratio balance: min={rb['min']:.2f} "
          f"max={rb['max']:.2f} spread={rb['spread']:.2f} CV={rb['cv']:.3f}):")
    lane_reports, all_hard = build_lane_reports(buckets, L7, L5, rc_i5, mg, mb, md, w34)
    for lr in lane_reports:
        cb = "PASS" if (lr["i7"]["hard_ok"] and lr["i5"]["hard_ok"]) else "FAIL"
        ids = []
        for b in buckets[lr["lane"] - 1]:
            ids.extend(m["id"] for m in b["members"])
        print(f"     Lane {lr['lane']}: {lr['n_libraries']} samples, ratio={lr['ratio_sum']:.2f}, "
              f"colorbal={cb}  ids: {','.join(ids)}")
    print(f"   Color balance all lanes pass: {'YES' if all_hard else 'NO'}")
    print()


# ============================================================================
# Commands
# ============================================================================

def cmd_check(args):
    items = read_table(args.input)
    validate_indexes(items)

    def evaluate_and_print(sub_items, label="ALL"):
        L7_req = args.i7_cycles if args.i7_cycles is not None else args.cycles
        L5_req = args.i5_cycles if args.i5_cycles is not None else args.cycles
        L7 = effective_cycles(sub_items, "i7", L7_req)
        L5 = effective_cycles(sub_items, "i5", L5_req)
        i7 = [r["i7"][:L7] for r in sub_items]
        i5 = [(revcomp(r["i5"]) if args.rc_i5 else r["i5"])[:L5] for r in sub_items]
        ratios = [r["ratio"] for r in sub_items]

        # barcode duplicate report (full strings, demux-correct)
        full_pairs = [(r["i7"], (revcomp(r["i5"]) if args.rc_i5 else r["i5"])) for r in sub_items]
        cnt = Counter(full_pairs)
        dups = {p: n for p, n in cnt.items() if n > 1}
        if dups:
            print(color_text(f"  !! {len(dups)} duplicated barcode(s) in this lane "
                             f"-> demultiplexing will FAIL:", "red", True))
            id_by_pair = defaultdict(list)
            for r, p in zip(sub_items, full_pairs):
                id_by_pair[p].append(r["id"])
            for p, n in dups.items():
                print(f"     {p[0]}+{p[1]} x{n}: {', '.join(id_by_pair[p])}")
        combined = [a + b for a, b in zip(i7, i5)]
        mh = min_pairwise_hamming(combined)
        if mh is not None:
            note = "  (collision risk!)" if mh <= 2 else ""
            print(f"  min pairwise Hamming (i7+i5, {L7+L5} bp): {mh}{note}")

        rep7 = evaluate_color_balance_strict_weighted(i7, ratios, L7, args.min_green, args.min_blue, args.max_dark, args.w34)
        rep5 = evaluate_color_balance_strict_weighted(i5, ratios, L5, args.min_green, args.min_blue, args.max_dark, args.w34)
        report = {
            "lanes": 1, "cycles": {"i7": L7, "i5": L5},
            "lane_reports": [{
                "lane": label, "n_libraries": len(sub_items),
                "ratio_sum": sum(ratios), "i7": rep7, "i5": rep5,
                "types": {}, "min_hamming": mh,
            }],
            "all_lanes_hard_pass": (rep7["hard_ok"] and rep5["hard_ok"]),
        }
        print_group_human(report, args.min_green, args.min_blue, args.max_dark, args.rc_i5, "")

    def lane_of(r):
        return str(r.get("lane", "")).strip()

    # A real lane label is anything other than blank or the "ALL" default (which marks
    # "no lane given"). Treat blank/ALL consistently here and in the label set below.
    real = {"" , "ALL"}
    has_lane = any(lane_of(r) not in real for r in items)
    if has_lane:
        labels = sorted({lane_of(r) for r in items if lane_of(r) not in real})
        unassigned = [r for r in items if lane_of(r) in real]
        if unassigned:
            ids = ", ".join(r["id"] for r in unassigned)
            print(color_text(f"⚠️  {len(unassigned)} row(s) have no lane and are skipped in the "
                             f"per-lane check (assign a lane or remove the column): {ids}",
                             "yellow", True))
            print()
        for lbl in labels:
            sub = [r for r in items if lane_of(r) == lbl]
            print(f"=== CHECK for lane {lbl} ===")
            evaluate_and_print(sub, lbl)
            print()
    else:
        print("=== CHECK for all libraries (single pool) ===")
        evaluate_and_print(items, "ALL")


# ============================================================================
# Freeze mode (--into): honor existing lane column + no-index ballast
# ============================================================================

def _validate_freeze(items) -> None:
    seen = set()
    for r in items:
        if r["id"] in seen:
            raise ValueError(f"Duplicate sample id '{r['id']}' in input (ids must be unique).")
        seen.add(r["id"])
        for k in ("i7", "i5"):
            for ch in r[k]:
                if ch not in "ACGTN":
                    raise ValueError(f"{r['id']} {k} contains invalid base '{ch}'.")


def _place_free_units(buckets, free_units, k, budget=300000):
    """Place free units into k buckets with NO barcode conflict, honoring units already in the
    buckets (frozen). Uses complete backtracking (most-constrained barcode first; try lightest
    non-conflicting lane first) so it does not falsely fail when a conflict-free placement
    exists; falls back to greedy if the step budget is exceeded. Returns (ok, offending_unit)."""
    mult = Counter()
    for b in buckets:
        for u in b:
            for p in u["full_pairs"]:
                mult[p] += 1
    for u in free_units:
        for p in u["full_pairs"]:
            mult[p] += 1
    order = sorted(free_units,
                   key=lambda u: (-max((mult[p] for p in u["full_pairs"]), default=0), -u["ratio"]))

    def cands_for(u):
        out = []
        for li in range(k):
            seen = {p for uu in buckets[li] for p in uu["full_pairs"]}
            if not any(p in seen for p in u["full_pairs"]):
                out.append(li)
        out.sort(key=lambda li: sum(uu["ratio"] for uu in buckets[li]))  # lightest first
        return out

    steps = [0]

    class _Budget(Exception):
        pass

    def bt(j):
        if j == len(order):
            return True
        steps[0] += 1
        if steps[0] > budget:
            raise _Budget()
        u = order[j]
        cands = cands_for(u)
        if not cands:
            return False
        for li in cands:
            buckets[li].append(u)
            if bt(j + 1):
                return True
            buckets[li].pop()
        return False

    try:
        if bt(0):
            return True, None
        # provably no conflict-free placement: name an over-subscribed barcode's unit
        for u in order:
            if not cands_for(u):
                return False, u
        return False, order[0]
    except _Budget:
        # budget exceeded: discard partial backtracking state, then best-effort greedy
        free_ids = {id(u) for u in free_units}
        for b in buckets:
            b[:] = [u for u in b if id(u) not in free_ids]
        for u in order:
            cands = cands_for(u)
            if not cands:
                return False, u
            buckets[cands[0]].append(u)
        return True, None


def _group_freeze(args, items):
    """--into mode: rows that already have a lane are FIXED; blank-lane rows are distributed
    into the --into lanes; rows pinned to lanes not in --into are ignored. No-index rows are
    ratio-only ballast (counted in balance, excluded from conflict/color); color is optimized
    only for lanes whose members are all fully indexed."""
    try:
        into = sorted(dict.fromkeys(int(x) for x in str(args.into).replace(" ", "").split(",") if x != ""))
    except ValueError:
        raise ValueError("--into must be comma-separated lane numbers, e.g. --into 1,2,3")
    if not into:
        raise ValueError("--into requires at least one lane.")
    pos = {lane: i for i, lane in enumerate(into)}
    k = len(into)

    def lane_of(it):
        return str(it.get("lane", "")).strip()

    free, frozen, dropped = [], [], []
    for it in items:
        l = lane_of(it)
        if l in ("", "ALL"):
            free.append(it)
        else:
            try:
                ln = int(l)
            except ValueError:
                dropped.append(it); continue
            (frozen if ln in pos else dropped).append(it)

    kept = free + frozen
    # Validate ALL rows (incl. ignored lanes): ids must be globally unique, otherwise the
    # id->lane write-back would silently move an ignored row onto a kept row's lane.
    _validate_freeze(items)
    bad = [it["id"] for it in free if not (it["i7"] and it["i5"])]
    if bad:
        raise ValueError("Unassigned (blank-lane) samples without a full index cannot be "
                         "distributed; give them a lane or an index: " + ", ".join(bad))

    indexed = [it for it in kept if it["i7"] and it["i5"]]
    L7_req = args.i7_cycles if args.i7_cycles is not None else args.cycles
    L5_req = args.i5_cycles if args.i5_cycles is not None else args.cycles
    L7 = effective_cycles(indexed, "i7", L7_req)
    L5 = effective_cycles(indexed, "i5", L5_req)
    w7 = make_weights_special(L7, args.w34)
    w5 = make_weights_special(L5, args.w34)
    mg, mb, md = args.min_green, args.min_blue, args.max_dark

    units, internal_conflicts = build_units(kept, args.rc_i5)
    finalize_units(units, L7, L5, args.rc_i5)
    if internal_conflicts:
        print(color_text("=== UNFIXABLE: bundle-internal barcode collisions ===", "red", True))
        for uid, members, dups in internal_conflicts:
            for p in dups:
                print(f"  bundle '{uid}': {p[0]}+{p[1]} repeated internally")
        print("Resolve these (re-index or drop a colliding member) and re-run.")
        return

    # determine fixed lane per unit from its members' lane (blank -> free)
    for u in units:
        lset = set()
        for m in u["members"]:
            ml = str(m.get("lane", "")).strip()
            lset.add("" if ml in ("", "ALL") else ml)
        if lset == {""}:
            u["fixed_lane"] = None
        elif len(lset) == 1:
            u["fixed_lane"] = int(next(iter(lset)))
        else:
            raise ValueError(f"'{u['uid']}' spans inconsistent lanes {lset}; a bundle must be one lane.")

    buckets = [[] for _ in into]
    free_units, locked = [], set()
    for u in units:
        if u["fixed_lane"] is None:
            free_units.append(u)
        else:
            buckets[pos[u["fixed_lane"]]].append(u)
            locked.add(u["uid"])

    n_ball = sum(1 for u in units if u.get("is_ballast"))
    print("=== Lane Assignment — freeze mode ===")
    print(f"Distribute blank-lane samples into lanes: {into}")
    print(f"Kept: {len(indexed)} indexed sample(s) + {n_ball} no-index ballast unit(s);  "
          f"frozen={len(locked)} unit(s), to-distribute={len(free_units)} unit(s).")
    if dropped:
        dl = sorted({lane_of(it) for it in dropped})
        print(f"Ignored (lane not in --into): {len(dropped)} sample(s) in lane(s) {dl}.")
    print()

    ok, bad_u = _place_free_units(buckets, free_units, k)
    if not ok:
        print(color_text(f"❌ INFEASIBLE: cannot place '{bad_u['uid']}' ({bad_u['full_pairs']}) "
                         f"conflict-free into any of lanes {into} — its barcode is over-subscribed "
                         f"for {k} lane(s). Add a lane to --into or drop a colliding sample.", "red", True))
        return

    # Freeze-mode priority matches the pooling intent: conflict (hard) >> read-count balance
    # (ENSURE) >> color (best-effort). Balance must dominate color here, otherwise the
    # optimizer would "hide" hard-to-color samples in pool-covered lanes (which carry no color
    # cost) at the expense of read balance. (Differs from default group mode, where color leads.)
    buckets = optimize_units(buckets, L7, L5, args.rc_i5, args.iters, args.seed, mg, mb, md, w7, w5,
                             c12_weight=1e3, balance_weight=1e7,
                             allow_moves=True, locked_uids=locked)

    _report_freeze(args, items, into, buckets, dropped, L7, L5, w7, w5)


def _report_freeze(args, items, into, buckets, dropped, L7, L5, w7, w5):
    mg, mb, md = args.min_green, args.min_blue, args.max_dark
    total = sum(u["ratio"] for b in buckets for u in b)
    target = total / len(into) if into else 0.0

    print("=== Final Conflict Check (indexed samples only) ===")
    conflict = False
    for i, lane in enumerate(into):
        pairs = [p for u in buckets[i] for p in u["full_pairs"]]
        dups = {p for p, n in Counter(pairs).items() if n > 1}
        if dups:
            conflict = True
            print(color_text(f"❌ [CRITICAL] Lane {lane} duplicate barcode(s): {dups}", "red", True))
        else:
            print(f"✅ Lane {lane} unique check passed.")
    print("✨ All distributed lanes free of barcode conflicts.\n" if not conflict
          else color_text("\n!! Conflict detected — DO NOT sequence !!\n", "red", True))

    print(f"=== Lanes (target ~{target:.0f} ratio/lane) ===")
    for i, lane in enumerate(into):
        b = buckets[i]
        rsum = sum(u["ratio"] for u in b)
        has_ball = any(u.get("is_ballast") for u in b)
        n_s = sum(u["nsamp"] for u in b)
        print(color_text(f"\nLane {lane}: ratio={rsum:.0f}  ({n_s} samples, dev {rsum-target:+.0f})"
                         f"{'  [pool-covered]' if has_ball else ''}", "magenta", True))
        for u in sorted(b, key=lambda u: (0 if u.get("is_ballast") else 1, -u["ratio"])):
            kind = "ballast" if u.get("is_ballast") else ("bundle" if u["is_bundle"] else "sample")
            tag = " <-frozen" if u["fixed_lane"] is not None else ""
            idx = "" if u.get("is_ballast") else f"  {u['members'][0]['i7']}+{u['members'][0]['i5']}"
            print(f"   [{kind:>7}] {u['uid']:<32} ratio={u['ratio']:>5.0f}{idx}{tag}")
        if has_ball:
            print("   color: N/A (lane contains a complex pool / ballast)")
        else:
            i7s = [x for u in b for x in u["m_i7"]]; i5s = [x for u in b for x in u["m_i5"]]
            rr = [x for u in b for x in u["m_ratio"]]
            r7 = evaluate_color_balance_strict_weighted(i7s, rr, L7, mg, mb, md, args.w34)
            r5 = evaluate_color_balance_strict_weighted(i5s, rr, L5, mg, mb, md, args.w34)
            i5lab = "i5(rc)" if args.rc_i5 else "i5    "
            print(f"   color i7    : C1-2 {'PASS' if r7['hard_ok'] else 'FAIL'}  wpen(>=3)={r7['weighted_penalty']:.3f}")
            print(f"   color {i5lab}: C1-2 {'PASS' if r5['hard_ok'] else 'FAIL'}  wpen(>=3)={r5['weighted_penalty']:.3f}")
    if args.rc_i5:
        print(color_text("\n  (i5 evaluated as reverse-complement)", "cyan"))

    rb = ratio_balance_stats(buckets)
    print(color_text(f"\nRatio balance across {len(into)} distributed lanes: min={rb['min']:.0f} "
                     f"max={rb['max']:.0f} spread={rb['spread']:.0f} CV={rb['cv']:.3f}", "cyan"))

    if args.out and conflict:
        print(color_text(f"⚠️  Refusing to write {args.out}: conflicts present.", "red", True))
    elif args.out:
        _write_freeze(args.out, items, into, buckets)


def _write_freeze(out_path, items, into, buckets):
    """Write the full table (frozen + newly-assigned + ignored) with the lane column filled."""
    lane_of_id = {}
    for i, lane in enumerate(into):
        for u in buckets[i]:
            for m in u["members"]:
                lane_of_id[m["id"]] = lane
    p = Path(out_path); p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["id", "i7", "i5", "ratio", "lane", "type", "bundle"])
        w.writeheader()
        for r in items:
            lane = lane_of_id.get(r["id"], str(r.get("lane", "")).strip())  # frozen/ignored keep theirs
            w.writerow({"id": r["id"], "i7": r["i7"], "i5": r["i5"], "ratio": r["ratio"],
                        "lane": lane, "type": r.get("type", "ALL"), "bundle": r.get("bundle", "")})
    print(f"Assignment written to: {out_path}")


def cmd_group(args):
    items = read_table(args.input)

    if getattr(args, "into", "") and getattr(args, "constraints", ""):
        raise ValueError("--constraints and --into are different assignment mechanisms and "
                         "cannot be combined; use one or the other.")

    if getattr(args, "into", ""):
        _group_freeze(args, items)   # honor existing lane column + no-index ballast
        return

    validate_indexes(items)

    L7_req = args.i7_cycles if args.i7_cycles is not None else args.cycles
    L5_req = args.i5_cycles if args.i5_cycles is not None else args.cycles
    L7 = effective_cycles(items, "i7", L7_req)
    L5 = effective_cycles(items, "i5", L5_req)
    w7 = make_weights_special(L7, args.w34)
    w5 = make_weights_special(L5, args.w34)
    mg, mb, md = args.min_green, args.min_blue, args.max_dark

    units, internal_conflicts = build_units(items, args.rc_i5)
    finalize_units(units, L7, L5, args.rc_i5)

    # ---- Unfixable internal bundle conflicts: report and stop ----
    if internal_conflicts:
        print(color_text("=== UNFIXABLE: bundle-internal barcode collisions ===", "red", True))
        print("These pre-pooled bundles contain samples sharing the SAME barcode. No lane")
        print("assignment can demultiplex them — the pool itself must be fixed:")
        for uid, members, dups in internal_conflicts:
            ids_by_pair = defaultdict(list)
            for m in members:
                p = (m["i7"], revcomp(m["i5"]) if args.rc_i5 else m["i5"])
                ids_by_pair[p].append(m["id"])
            for p in dups:
                print(f"  bundle '{uid}': {p[0]}+{p[1]} shared by {', '.join(ids_by_pair[p])}")
        print("\nResolve these (re-index or drop a colliding member) and re-run.")
        return

    # ---- Lane count: --constraints (if given) takes priority and sets the lane count ----
    constraints = None
    if args.constraints:
        cdata = load_constraints(args.constraints)
        lane_ids, lane_quotas, lane_must = normalize_constraints(cdata)
        if args.lanes and args.lanes != len(lane_ids):
            print(color_text(f"[INFO] --lanes={args.lanes} overridden by --constraints "
                             f"({len(lane_ids)} lanes).", "cyan"))
        lanes = len(lane_ids)
        constraints = (lane_ids, lane_quotas, lane_must)
    else:
        if not args.lanes or args.lanes < 1:
            raise ValueError("--lanes is required and must be >= 1 (or provide --constraints).")
        lanes = args.lanes

    n_units = len(units)
    n_bundles = sum(1 for u in units if u["is_bundle"])
    idx_to_units = build_index_map(units)
    adj = build_adjacency(n_units, idx_to_units)
    dup_groups = {p: us for p, us in idx_to_units.items() if len(us) > 1}
    max_mult = max_multiplicity(idx_to_units)

    print("=== Feasibility Analysis ===")
    print(f"Libraries: {len(items)} samples in {n_units} pooling unit(s) "
          f"({n_bundles} bundle(s), {n_units - n_bundles} singletons).")
    if dup_groups:
        print(f"Shared barcodes across units: {len(dup_groups)} "
              f"(max multiplicity = {max_mult} -> needs >= {max_mult} lanes).")
        for p, us in sorted(dup_groups.items(), key=lambda kv: -len(kv[1])):
            who = ", ".join(units[i]["uid"] for i in us)
            print(f"   {p[0]}+{p[1]} x{len(us)}: {who}")
    else:
        print("Shared barcodes across units: none.")
    print(f"Requested lanes: {lanes}\n")

    # ---- Constraints path (highest priority): place per user quotas/must_include, then
    #      let the automatic engine optimize WITHIN that structure (conflict -> C1/C2 ->
    #      ratio balance -> >=C3 color), without violating the constraints. ----
    if constraints:
        _group_with_constraints(args, items, units, lanes, constraints, L7, L5, w7, w5, dup_groups)
        return

    coloring, exact = color_units(adj, lanes)
    feasible = coloring is not None

    # ---------------- INFEASIBLE (or UNDETERMINED) ----------------
    if not feasible:
        min_lanes, mexact = min_lanes_needed(adj, idx_to_units)
        if exact:
            # The exact search proved no conflict-free layout exists for the requested lanes.
            print(color_text(f"❌ INFEASIBLE: {len(items)} samples cannot fit into "
                             f"{lanes} lane(s) without barcode conflicts.", "red", True))
        else:
            # The exact search hit its step budget and the heuristic also failed: this is
            # "unknown", NOT a proof of infeasibility. Report honestly; the analysis below
            # (min lanes, removal plans) is best-effort.
            print(color_text(f"⚠️  UNDETERMINED: could not find a conflict-free layout for "
                             f"{lanes} lane(s) within the search budget — it may still be "
                             f"feasible. Best-effort analysis follows.", "yellow", True))
        approx = "" if mexact else " (estimate)"
        print(color_text(f"   (1) Minimum lanes required: {min_lanes}{approx}", "yellow", True))

        if args.auto_expand:
            print(color_text(f"\n--auto-expand set: solving at {min_lanes} lanes.\n", "green", True))
            buckets = assign_and_optimize(units, min_lanes, L7, L5, args.rc_i5,
                                          args.iters, args.seed, mg, mb, md, w7, w5,
                                          args.c12_weight, args.balance_weight)
            _finish_feasible(args, buckets, min_lanes, L7, L5, w7, w5, dup_groups, expanded_from=lanes)
            return

        print(color_text(f"   (2) Removal options to fit into {lanes} lane(s):", "yellow", True))
        plans = generate_removal_plans(units, lanes, max_plans=args.max_plans)
        if not plans:
            print("   (could not derive a removal plan automatically)")
            return
        for k, rem in enumerate(plans, start=1):
            removed_units = [units[i] for i in rem]
            removed_set = set(rem)
            kept = [units[i] for i in range(n_units) if i not in removed_set]
            buckets = assign_and_optimize(kept, lanes, L7, L5, args.rc_i5,
                                          min(args.iters, 8000), args.seed,
                                          mg, mb, md, w7, w5, args.c12_weight, args.balance_weight)
            if buckets is None:
                print(f"--- Removal Plan {k}: (failed to lay out remaining samples) ---")
                continue
            print_plan(k, removed_units, buckets, L7, L5, args.rc_i5, mg, mb, md, args.w34)
            if args.out:
                _write_assignment(args.out, items, buckets, removed_units, suffix=f"_plan{k}")
        if args.out:
            print(color_text(f"Per-plan assignments written next to: {args.out} (suffix _planN).", "cyan"))
        return

    # ---------------- FEASIBLE ----------------
    print(color_text(f"✅ FEASIBLE: all {len(items)} samples fit into {lanes} lane(s) "
                     f"with no barcode conflicts.", "green", True))
    if dup_groups:
        print(color_text(f"   {len(dup_groups)} shared-barcode group(s) auto-separated across lanes.", "green"))
    print()
    buckets = assign_and_optimize(units, lanes, L7, L5, args.rc_i5,
                                  args.iters, args.seed, mg, mb, md, w7, w5,
                                  args.c12_weight, args.balance_weight)
    _finish_feasible(args, buckets, lanes, L7, L5, w7, w5, dup_groups)


def _group_with_constraints(args, items, units, lanes, constraints, L7, L5, w7, w5, dup_groups):
    """Constraints take priority: place samples per user quotas/must_include, then optimize
    color + ratio balance WITHIN that structure (quota-preserving swaps, must_include locked)."""
    lane_ids, lane_quotas, lane_must = constraints
    mg, mb, md = args.min_green, args.min_blue, args.max_dark

    # resolve must_include ids (sample id OR bundle id) -> unit index
    id_to_unit = {}
    for k, u in enumerate(units):
        id_to_unit[u["uid"]] = k
        for m in u["members"]:
            id_to_unit[m["id"]] = k

    n_quota = sum(sum(q.values()) for q in lane_quotas.values())
    n_must = sum(len(s) for s in lane_must.values())
    print(color_text(f"Applying --constraints (priority): {lanes} lanes, {n_quota} quota slots, "
                     f"{n_must} must_include pin(s).", "cyan", True))

    check_quota_feasibility(items, lane_ids, lane_quotas)
    buckets = place_units_with_quotas(units, lane_ids, lane_quotas, lane_must, id_to_unit)

    locked = {units[id_to_unit[sid]]["uid"] for lid in lane_ids for sid in lane_must[lid]}

    buckets = optimize_units(buckets, L7, L5, args.rc_i5, args.iters, args.seed,
                             mg, mb, md, w7, w5,
                             c12_weight=args.c12_weight, balance_weight=args.balance_weight,
                             locked_uids=locked, quota_lock=True)

    print(color_text("Quotas / must_include satisfied — now checking barcode conflicts...", "cyan"))
    if dup_groups:
        print(color_text(f"   (duplicate barcodes are separated where the quotas allow)", "cyan"))
    print()
    _finish_feasible(args, buckets, lanes, L7, L5, w7, w5, dup_groups)


def _finish_feasible(args, buckets, lanes, L7, L5, w7, w5, dup_groups, expanded_from=None):
    mg, mb, md = args.min_green, args.min_blue, args.max_dark
    lane_reports, all_hard = build_lane_reports(buckets, L7, L5, args.rc_i5, mg, mb, md, args.w34)
    rb = ratio_balance_stats(buckets)

    # final safety: confirm no barcode collisions anywhere
    print("=== Final Conflict Check ===")
    conflict_found = False
    for i, bucket in enumerate(buckets, start=1):
        pairs = _lane_pairs(bucket)
        dups = {p for p, n in Counter(pairs).items() if n > 1}
        if dups:
            conflict_found = True
            ids_by_pair = defaultdict(list)
            for u in bucket:
                for m, p in zip(u["members"], u["full_pairs"]):
                    if p in dups:
                        ids_by_pair[p].append(m["id"])
            detail = "; ".join(f"{p[0]}+{p[1]} -> {', '.join(ids_by_pair[p])}" for p in dups)
            print(color_text(f"❌ [CRITICAL] Lane {i} duplicate barcode(s): {detail}", "red", True))
        else:
            print(f"✅ Lane {i} unique check passed.")
    print("✨ All lanes free of barcode conflicts.\n" if not conflict_found
          else color_text("\n!! Conflict detected — DO NOT sequence this layout !!\n", "red", True))

    report = {
        "lanes": lanes, "cycles": {"i7": L7, "i5": L5},
        "lane_reports": lane_reports, "all_lanes_hard_pass": all_hard,
        "ratio_balance": rb,
    }
    # Refuse to write a colliding assignment: the final safety check exists precisely to
    # stop a dangerous layout from being saved (should never trigger on the normal path).
    write_path = args.out if (args.out and not conflict_found) else ""
    print_group_human(report, mg, mb, md, args.rc_i5, out_path=write_path)
    if expanded_from is not None:
        print(color_text(f"(Note: requested {expanded_from} lanes was infeasible; "
                         f"used {lanes} lanes instead.)", "yellow", True))

    if args.out and conflict_found:
        print(color_text(f"⚠️  Refusing to write {args.out}: the assignment contains barcode "
                         f"conflicts (see above). No file written.", "red", True))
    elif args.out:
        _write_assignment(args.out, read_table(args.input), buckets, removed_units=None)


def _write_assignment(out_path, original_items, buckets, removed_units=None, suffix=""):
    """Write id,i7,i5,ratio,lane,type,bundle preserving input order; dropped samples excluded."""
    p = Path(out_path)
    if suffix:
        p = p.with_name(p.stem + suffix + p.suffix)
    p.parent.mkdir(parents=True, exist_ok=True)

    lane_of = {}
    for lane_id, buck in enumerate(buckets, start=1):
        for u in buck:
            for m in u["members"]:
                lane_of[m["id"]] = lane_id

    with p.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["id", "i7", "i5", "ratio", "lane", "type", "bundle"])
        w.writeheader()
        for r in original_items:
            if r["id"] not in lane_of:
                continue  # removed by a plan
            w.writerow({"id": r["id"], "i7": r["i7"], "i5": r["i5"], "ratio": r["ratio"],
                        "lane": lane_of[r["id"]], "type": r.get("type", "ALL"),
                        "bundle": r.get("bundle", "")})


# ============================================================================
# CLI
# ============================================================================

def build_parser():
    p = argparse.ArgumentParser(
        description="Index color-balance checker & lane balancer v3 "
                    "(auto-separates duplicate barcodes; atomic bundles; feasibility + removal plans).")
    sub = p.add_subparsers(dest="cmd", required=True)

    def add_common(a):
        a.add_argument("input", help="CSV/TSV/XLSX with headers id,i7,i5 and optional ratio,type,lane,bundle")
        a.add_argument("--cycles", type=int, default=8, help="Cycles to evaluate for i7 and i5 (default 8)")
        a.add_argument("--i7-cycles", type=int, default=None, help="Override cycles for i7")
        a.add_argument("--i5-cycles", type=int, default=None, help="Override cycles for i5")
        a.add_argument("--min-green", type=float, default=0.25, help="Min green fraction per cycle (default 0.25)")
        a.add_argument("--min-blue", type=float, default=0.05, help="Min blue fraction per cycle (default 0.05, strict '>')")
        a.add_argument("--max-dark", type=float, default=0.40, help="Max dark fraction per cycle (default 0.40)")
        a.add_argument("--w34", type=float, default=0.60, help="Weight share for cycles 3/4 (default 0.60)")
        a.add_argument("--rc-i5", action="store_true", help="Reverse-complement i5 before evaluating (default OFF)")

    pc = sub.add_parser("check", help="Check color balance + barcode duplicates per lane.")
    add_common(pc)
    pc.set_defaults(func=cmd_check)

    pg = sub.add_parser("group", help="Assign libraries into lanes; auto-separate duplicate barcodes; "
                                      "honor bundles; report feasibility / removal plans.")
    add_common(pg)
    pg.add_argument("--lanes", type=int, default=0,
                    help="Number of lanes to assign into (required unless --constraints or --into is given)")
    pg.add_argument("--constraints", type=str, default="",
                    help="JSON/YAML file of per-lane quotas + must_include. Takes PRIORITY and sets "
                         "the lane count; the automatic engine then optimizes within it.")
    pg.add_argument("--into", type=str, default="",
                    help="Freeze mode, e.g. --into 1,2,3 : rows that already have a lane are FIXED; "
                         "blank-lane rows are distributed into these lanes (rows pinned to other lanes "
                         "are ignored). No-index rows act as ratio-only ballast; color is optimized "
                         "only for lanes whose samples are all fully indexed.")
    pg.add_argument("--iters", type=int, default=20000, help="Swap/move iterations (default 20000)")
    pg.add_argument("--seed", type=int, default=42, help="Random seed (default 42)")
    pg.add_argument("--c12-weight", type=float, default=1e6, help="Priority weight for cycle 1/2 (default 1e6)")
    pg.add_argument("--balance-weight", type=float, default=10.0,
                    help="Weight for balancing per-lane ratio sums. Priority is "
                         "conflict >> cycle1/2 color >> ratio balance >> cycle>=3 color. "
                         "Set 0 for v2-style behavior (fixed round-robin lane sizes, color only). "
                         "(default 10.0)")
    pg.add_argument("--max-plans", type=int, default=3, help="Max removal plans to show when infeasible (default 3)")
    pg.add_argument("--auto-expand", action="store_true",
                    help="If requested lanes are infeasible, automatically solve at the minimum feasible lanes")
    pg.add_argument("--out", type=str, default="", help="Write lane assignment CSV (id,i7,i5,ratio,lane,type,bundle)")
    pg.set_defaults(func=cmd_group)

    return p


def main():
    args = build_parser().parse_args()
    args.func(args)


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Index color-balance checker & lane balancer for Illumina 2-color (NovaSeq X/X Plus).

Rules:
  T: green only
  C: green + blue
  A: blue only
  G: dark

Feasibility-first (cycles 1/2):
  - If there exists an assignment where EVERY lane's cycles 1 & 2 meet thresholds
    (per index type i7/i5 on lane-aggregate), the optimizer will find and prefer it.
  - If that is impossible for the given library set, the optimizer minimizes the total
    deviation from the thresholds on cycles 1 & 2 (lexicographic objective), while
    still optimizing cycles ≥3 with weights.

Weighted optimization (soft, for cycles ≥3):
  - By default, cycles 3/4 share 60% of the total weight (evenly),
    and cycles 5..N share the remaining 40% (evenly).
  - Cycles 1 and 2 have zero weight because they are prioritized via feasibility-first.
  - You can adjust the 60% via --w34 (0~1).

Input CSV/TSV must have headers: id,i7,i5
i5 is NOT reverse-complemented by default. Use --rc-i5 if needed.

easy run:
  python index_balance.py group library.csv --lanes 2 --out lanes.csv 
  # library.csv is a csv/tsv/xlsx with headers: id,i7,i5,[ratio=1,type=ALL]

  python index_balance.py check lanes.csv 
  # lanes.csv is a csv/tsv/xlsx with headers: id,i7,i5,[ratio=1,lane=ALL]

Outputs (human-readable):
  - check: overall info + per-cycle color fractions and pass/fail for i7 and i5
  - group: lane summary + per-cycle details for each lane (i7 and i5)
"""

import argparse
import csv
import random
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import openpyxl

BASES = ("A", "C", "G", "T")
BIG_PENALTY = 1e9  # kept for backward-compat, no longer used as a hard plateau

# -------------------- IO & utils --------------------

def revcomp(seq: str) -> str:
    comp = str.maketrans("ACGTNacgtn", "TGCANtgcan")
    return seq.translate(comp)[::-1]


def read_table(path: str, col_name: str) -> List[Dict[str, str]]:
    """
    Read CSV/TSV or XLSX with headers id,i7,i5[,ratio][,type/,lane]; handle CSV BOM; trim header whitespace.
    - CSV/TSV: auto delimiter by suffix (.csv => ',', otherwise '\t'), encoding='utf-8-sig'
    - XLSX/XLSM: use openpyxl, read the first sheet's first non-empty row as headers
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    suffix = p.suffix.lower()
    items: List[Dict[str, str]] = []

    # ---- Excel path ----
    if suffix in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise ImportError(
                "Reading XLSX/XLSM requires 'openpyxl'. Install with: pip install openpyxl"
            )
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb.active  # 使用第一个工作表

        # 找到第一行非空作为表头
        header = None
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            vals = [("" if v is None else str(v)).strip() for v in row]
            if any(v != "" for v in vals):
                header = vals
                break
        if not header:
            raise ValueError("No header row found in the Excel sheet.")

        # 规范化表头（去空白、去BOM、小写）
        header = [h.replace("\ufeff", "").strip() for h in header]
        header_lower = [h.lower() for h in header]
        # 建立列名->索引映射
        colmap = {name: idx for idx, name in enumerate(header_lower)}

        required = {"id", "i7", "i5"}
        missing = required - set(colmap.keys())
        if missing:
            raise ValueError(f"Missing required columns: {missing}. Found: {header}")

        # 继续读数据行
        for row in ws.iter_rows(min_row=ws.min_row + 1, values_only=True):
            if row is None:
                continue
            def get(col):
                idx = colmap[col]
                v = "" if idx >= len(row) or row[idx] is None else str(row[idx])
                return v.strip()
            rec = {"id": get("id"), "i7": get("i7").upper(), "i5": get("i5").upper(), "ratio": float(get("ratio") or 1), col_name: (get(col_name) or "ALL")}
            # 跳过完全空行
            if not (rec["id"] or rec["i7"] or rec["i5"]):
                continue
            items.append(rec)
        wb.close()
        return items

    # ---- CSV/TSV path ----
    # 默认：.csv 用逗号，其他后缀（.tsv/.txt）用制表符
    delim = "," if suffix == ".csv" else "\t"
    with p.open(encoding="utf-8-sig") as f:  # 处理 BOM
        reader = csv.DictReader(f, delimiter=delim)
        # 规范化表头
        fieldnames = [fn.replace("\ufeff", "").strip() for fn in (reader.fieldnames or [])]
        reader.fieldnames = fieldnames
        required = {"id", "i7", "i5"}
        missing = required - set(fn.lower() for fn in fieldnames)
        if missing:
            raise ValueError(f"Missing required columns: {missing}. Found: {fieldnames}")
        for row in reader:
            items.append({
                "id": (row.get("id") or row.get("ID") or "").strip(),
                "i7": (row.get("i7") or row.get("I7") or "").strip().upper(),
                "i5": (row.get("i5") or row.get("I5") or "").strip().upper(),
                "ratio": float((row.get("ratio") or '1').strip()),
                col_name: (row.get(col_name) or "ALL").strip()
            })
    return items


def read_table_group(path: str) -> List[Dict[str, str]]:
    return read_table(path, "type")


def read_table_check(path: str) -> List[Dict[str, str]]:
    return read_table(path, "lane")


def validate_indexes(items: List[Dict[str, str]]) -> None:
    for r in items:
        for k in ("i7", "i5"):
            s = r[k]
            if not s:
                raise ValueError(f"{r['id']} has empty {k}.")
            for ch in s:
                if ch not in "ACGTN":
                    raise ValueError(f"{r['id']} {k} contains invalid base '{ch}'. Only A/C/G/T/N allowed.")


def effective_cycles(items: List[Dict[str, str]], which: str, requested: int) -> int:
    """Use min(requested, min_length_across_items)."""
    min_len = min(len(r[which]) for r in items) if items else 0
    return max(0, min(requested, min_len))


# -------------------- Lane constraints (types & must-include) --------------------
from pathlib import Path as _Path_for_constraints
import json as _json_for_constraints


def _try_load_yaml(path: str):
    try:
        import yaml  # optional
    except Exception:
        return None, "PyYAML not installed"
    with open(path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f), None


def load_constraints(path: str) -> dict:
    p = _Path_for_constraints(path)
    if not p.exists():
        raise FileNotFoundError(f"Constraints file not found: {path}")
    suf = p.suffix.lower()
    if suf == ".json":
        with open(p, "r", encoding="utf-8") as f:
            data = _json_for_constraints.load(f)
    elif suf in (".yml", ".yaml"):
        data, err = _try_load_yaml(str(p))
        if data is None:
            raise RuntimeError(f"YAML not available: {err}. Please provide JSON instead.")
    else:
        raise ValueError("Constraints must be .json or .yaml/.yml")
    if not isinstance(data, dict) or "lanes" not in data or not isinstance(data["lanes"], dict):
        raise ValueError("Constraints must have top-level 'lanes' mapping")
    return data


def normalize_constraints(data: dict):
    lanes_cfg = data["lanes"]
    lane_ids = sorted(int(k) for k in lanes_cfg.keys())
    lane_quotas = {}
    lane_must = {}
    for lid in lane_ids:
        cfg = lanes_cfg.get(str(lid), lanes_cfg.get(lid))
        quotas = cfg.get("quotas", {}) or {}
        must = cfg.get("must_include", []) or []
        if not isinstance(quotas, dict):
            raise ValueError(f"lane {lid}: 'quotas' must be a dict of type->count")
        if not all(isinstance(v, int) and v >= 0 for v in quotas.values()):
            raise ValueError(f"lane {lid}: quota counts must be non-negative integers")
        lane_quotas[lid] = {str(t): int(c) for t, c in quotas.items()}
        lane_must[lid] = set(str(s) for s in must)
    return lane_ids, lane_quotas, lane_must


def check_quota_feasibility(items: list, lane_ids, lane_quotas):
    from collections import Counter
    avail = Counter([r.get("type", "ALL") for r in items])
    need_total = {}
    for lid in lane_ids:
        for t, c in lane_quotas[lid].items():
            need_total[t] = need_total.get(t, 0) + c
    if sum(need_total.values()) != len(items):
        raise ValueError(f"Sum of all lane quotas ({sum(need_total.values())}) must equal number of input libraries ({len(items)}).")
    for t, need in need_total.items():
        if avail.get(t, 0) < need:
            raise ValueError(f"Not enough samples of type '{t}': need {need}, available {avail.get(t,0)}.")


def initial_buckets_with_quotas(items: list, lane_ids, lane_quotas, lane_must):
    from collections import defaultdict, deque
    id_map = {r["id"]: r for r in items}
    pools = defaultdict(list)
    for r in items:
        pools[r.get("type", "ALL")].append(r)
    for t in pools:
        pools[t].sort(key=lambda r: (r["i7"] + "_" + r["i5"], r["id"]))
        pools[t] = deque(pools[t])

    buckets = {lid: [] for lid in lane_ids}
    remaining = {lid: dict(q) for lid, q in lane_quotas.items()}

    # must_include 先放
    for lid in lane_ids:
        for sid in lane_must[lid]:
            if sid not in id_map:
                raise ValueError(f"lane {lid} must_include sample '{sid}' not found in input.")
            r = id_map[sid]
            tt = r.get("type", "ALL")
            if remaining[lid].get(tt, 0) <= 0:
                raise ValueError(f"lane {lid} has no remaining quota for type '{tt}' to place must_include '{sid}'.")
            buckets[lid].append(r)
            remaining[lid][tt] -= 1
            if r in pools[tt]:
                pools[tt].remove(r)

    # 再按配额填满
    for lid in lane_ids:
        for tt, cnt in list(remaining[lid].items()):
            for _ in range(cnt):
                if not pools[tt]:
                    raise ValueError(f"Insufficient samples to fulfill lane {lid} type '{tt}' quota.")
                buckets[lid].append(pools[tt].popleft())
            remaining[lid][tt] = 0

    return [buckets[lid] for lid in lane_ids]

# -------------------- Color math --------------------

def per_cycle_base_lists(seqs: List[str], ratios: List[float], L: int) -> List[Dict[str, float]]:
    """
    Collect bases per cycle (ignore N).
    weighted by ratios
    """
    cycles = [{'A': 0, 'C': 0, 'G': 0, 'T': 0} for _ in range(L)]

    for j in range(len(seqs)):
        s = seqs[j]
        r = ratios[j]

        for i in range(L):
            b = s[i]
            if b in BASES:
                cycles[i][b] += r
    
    for i in range(L):
        tot = sum(cycles[i].values())
        if tot == 0:
            continue
        for b in BASES:
            cycles[i][b] /= tot

    return cycles


def color_fractions_one_cycle(base_list: Dict[str, float]) -> Dict[str, float]:
    """
    2-color mapping:
      T: green only
      C: green + blue
      A: blue only
      G: dark
    Each base contributes up to 2 "signal units".
    """
    a, t, c, g = base_list['A'], base_list['T'], base_list['C'], base_list['G']

    tot_units = 2 * (a + t + c + g)
    if tot_units == 0:
        return {"green": 0.0, "blue": 0.0, "dark": 0.0}
    
    green = (t * 2.0 + c * 1.0) / tot_units
    blue  = (a * 2.0 + c * 1.0) / tot_units
    dark  = (g * 2.0) / tot_units
    return {"green": green, "blue": blue, "dark": dark}


def hard_cycles_ok_aggregate(
    seqs: List[str],
    ratios: List[float],
    L: int,
    min_green: float,
    min_blue: float,
    max_dark: float
) -> Tuple[bool, List[Optional[Dict[str, float]]]]:
    """
    Check hard constraints for cycle 1 and 2 on the aggregate (lane-level / batch-level).
    Returns (ok, [c1_color, c2_color]) where entries may be None if L<cycle.
    """
    blists = per_cycle_base_lists(seqs, ratios, L)
    colors_12: List[Optional[Dict[str, float]]] = [None, None]
    for idx in (0, 1):  # cycles 1 and 2
        if L > idx:
            col = color_fractions_one_cycle(blists[idx])
            colors_12[idx] = col
            ok = (col["green"] >= min_green) and (col["blue"] > min_blue) and (col["dark"] <= max_dark)
            if not ok:
                return False, colors_12
    return True, colors_12


def make_weights_special(L: int, w34: float) -> List[float]:
    """
    Build per-cycle weights for penalty (sum to 1 over cycles >=3; cycles 1-2 weight=0).
    - If L <= 2: all zeros.
    - If 3 <= L <= 4: distribute all weight uniformly over cycles 3..L.
    - If L > 4: cycles 3/4 share 'w34' (evenly), cycles 5..L share (1-w34) (evenly).
    """
    if L <= 2:
        return [0.0] * L

    # clamp w34
    if w34 < 0.0:
        w34 = 0.0
    if w34 > 1.0:
        w34 = 1.0

    w = [0.0] * L
    if 3 <= L <= 4:
        mass = 1.0
        n = L - 2
        for i in range(2, L):
            w[i] = mass / n
    else:
        w[2] = w34 / 2.0  # cycle 3
        w[3] = w34 / 2.0  # cycle 4
        n_tail = L - 4
        if n_tail > 0:
            tail_share = (1.0 - w34) / n_tail
            for i in range(4, L):
                w[i] = tail_share

    total = sum(w)
    if total > 0:
        w = [x / total for x in w]
    return w


def weighted_penalty(
    seqs: List[str],
    ratios: List[float],
    L: int,
    min_green: float,
    min_blue: float,
    max_dark: float,
    weights: List[float]
) -> float:
    """Sum_i weights[i]*vi, where vi is threshold violation at cycle i (cycles 1-2 weights should be 0)."""
    if L <= 0 or not seqs or not weights:
        return 0.0
    blists = per_cycle_base_lists(seqs, ratios, L)
    pen = 0.0
    for i in range(min(L, len(weights))):
        if weights[i] == 0.0:
            continue
        col = color_fractions_one_cycle(blists[i])
        v = 0.0
        v += max(0.0, (min_green - col["green"]))
        v += max(0.0, (min_blue  - col["blue"]))
        v += max(0.0, (col["dark"] - max_dark))
        pen += weights[i] * v
    return pen

# -------------------- NEW: cycles 1/2 deviation (for feasibility-first) --------------------

def c12_violation_aggregate(
    seqs: List[str],
    ratios: List[float],
    L: int,
    min_green: float,
    min_blue: float,
    max_dark: float,
    strict_blue: bool = True
) -> float:
    """
    Sum of violations for cycles 1 and 2 (aggregate over all sequences).
    If 'strict_blue' is True, mimic the hard check's 'blue > min_blue' by
    adding a tiny epsilon so that blue == min_blue counts as a tiny violation.
    """
    if not seqs or L <= 0:
        return 0.0
    EPS = 1e-9 if strict_blue else 0.0
    blists = per_cycle_base_lists(seqs, ratios, L)
    pen = 0.0
    for idx in (0, 1):  # cycles 1 and 2
        if L > idx:
            col = color_fractions_one_cycle(blists[idx])
            v = 0.0
            v += max(0.0, (min_green - col["green"]))
            v += max(0.0, (min_blue  - col["blue"]) + EPS)  # strict '>'
            v += max(0.0, (col["dark"] - max_dark))
            pen += v
    return pen


def evaluate_color_balance_strict_weighted(
    seqs: List[str],
    ratios: List[float],
    cycles: int,
    min_green: float,
    min_blue: float,
    max_dark: float,
    w34: float
) -> Dict[str, Any]:
    """
    For CHECK / per-lane report:
      - Enforce hard constraints on cycles 1 and 2 (boolean for reporting).
      - Report per-cycle pass/fail for 1..cycles.
      - Report weighted penalty over cycles >=3 using make_weights_special(cycles, w34).
    """
    # hard constraints on C1/C2 (for reporting)
    hard_ok, colors_12 = hard_cycles_ok_aggregate(seqs, ratios, cycles, min_green, min_blue, max_dark)

    # per-cycle info
    blists = per_cycle_base_lists(seqs, ratios, cycles)
    per_cycle: List[Dict[str, Any]] = []
    worst = None
    worst_score = -1.0
    for i, bl in enumerate(blists, start=1):
        col = color_fractions_one_cycle(bl)
        ok = (col["green"] >= min_green) and (col["blue"] > min_blue) and (col["dark"] <= max_dark)
        per_cycle.append({"cycle": i, "color": col, "ok": ok})
        # total violation for "worst cycle"
        v = max(0.0, (min_green - col["green"])) \
          + max(0.0, (min_blue  - col["blue"])) \
          + max(0.0, (col["dark"] - max_dark))
        if v > worst_score:
            worst_score = v
            worst = {"cycle": i, "color": col, "color_ok": (v == 0.0)}

    # mean color (informational)
    denom = cycles if cycles > 0 else 1
    mean_color = {
        "green": (sum(pc["color"]["green"] for pc in per_cycle) / denom) if per_cycle else 0.0,
        "blue":  (sum(pc["color"]["blue"]  for pc in per_cycle) / denom) if per_cycle else 0.0,
        "dark":  (sum(pc["color"]["dark"]  for pc in per_cycle) / denom) if per_cycle else 0.0
    }

    # weighted penalty over cycles >=3
    weights = make_weights_special(cycles, w34)
    wpen = weighted_penalty(seqs, ratios, cycles, min_green, min_blue, max_dark, weights)

    return {
        "cycles": cycles,
        "hard_ok": hard_ok,
        "hard_c12": colors_12,           # [c1_color, c2_color] or None
        "per_cycle": per_cycle,          # list of {cycle, color{...}, ok}
        "weighted_penalty": wpen,        # cycles >=3
        "weights": weights,              # transparency
        "mean_color": mean_color,
        "worst_cycle": worst
    }

# -------------------- Lane optimization (feasibility-first on C1/C2) --------------------

def lane_cost_with_hard(
    i7_seqs: List[str],
    i5_seqs: List[str],
    ratios: List[float],
    L7: int, L5: int,
    min_green: float, min_blue: float, max_dark: float,
    w7: List[float],
    w5: List[float],
    c12_weight: float = 1e6
) -> float:
    """
    Lexicographic/feasibility-first cost:
      1) Minimize sum of C1/C2 violations (i7 + i5).
      2) If (1) == 0, minimize weighted penalties on cycles ≥3 (i7 + i5).

    Total cost = c12_weight * (c12_violation_i7 + c12_violation_i5) + (pen7 + pen5)

    With sufficiently large c12_weight, any feasible assignment (C1/C2 violations == 0)
    is strictly preferred over any infeasible one, while avoiding BIG_PENALTY plateaus.
    """
    # primary term: C1/C2 violations (aggregate, lane-level)
    c12_pen7 = c12_violation_aggregate(i7_seqs, ratios, L7, min_green, min_blue, max_dark, strict_blue=True)
    c12_pen5 = c12_violation_aggregate(i5_seqs, ratios, L5, min_green, min_blue, max_dark, strict_blue=True)
    c12_total = c12_pen7 + c12_pen5

    # secondary term: ≥C3 weighted penalties
    pen7 = weighted_penalty(i7_seqs, ratios, L7, min_green, min_blue, max_dark, w7)
    pen5 = weighted_penalty(i5_seqs, ratios, L5, min_green, min_blue, max_dark, w5)

    return c12_weight * c12_total + (pen7 + pen5)


def initial_round_robin(items: List[Dict[str, str]], lanes: int) -> List[List[Dict[str, str]]]:
    buckets = [[] for _ in range(lanes)]
    keyed = sorted(items, key=lambda r: (r["i7"] + "_" + r["i5"]))
    for i, r in enumerate(keyed):
        buckets[i % lanes].append(r)
    return buckets


def optimize_assignment(
    buckets: List[List[Dict[str, str]]],
    L7: int, L5: int,
    rc_i5: bool,
    iters: int,
    seed: int,
    min_green: float, min_blue: float, max_dark: float,
    w7: List[float],
    w5: List[float],
    c12_weight: float = 1e6,
    lock_types: bool = False,
    type_field: str = 'type',
    locked_by_lane: dict | None = None
) -> List[List[Dict[str, str]]]:
    """
    Random swap optimizer with lexicographic cost:
      - Prioritize reducing C1/C2 total violations across lanes (i7+i5).
      - Once zero, further optimize ≥C3 weighted penalties.
    """
    random.seed(seed)
    lanes = len(buckets)

    def cost(buck: List[Dict[str, str]]) -> float:
        i7 = [r["i7"][:L7] for r in buck]
        i5s = [revcomp(r["i5"]) if rc_i5 else r["i5"] for r in buck]
        i5 = [s[:L5] for s in i5s]
        ratios = [r["ratio"] for r in buck]

        return lane_cost_with_hard(
            i7, i5, ratios, L7, L5,
            min_green, min_blue, max_dark,
            w7, w5,
            c12_weight=c12_weight
        )

    lane_scores = [cost(b) for b in buckets]
    total = sum(lane_scores)

    # 准备锁映射（lane_id -> {sample_id,...}）；无约束时为空字典
    locked_by_lane = locked_by_lane or {}

    for _ in range(iters):
        a = random.randrange(lanes)
        b = random.randrange(lanes)
        if a == b or not buckets[a] or not buckets[b]:
            continue

        if lock_types:
            # 在 a 里挑一个未锁定样本
            cand_a = [i for i, r in enumerate(buckets[a])
                      if r.get(type_field) is not None
                      and r["id"] not in locked_by_lane.get(a + 1, set())]
            if not cand_a:
                continue
            ia = random.choice(cand_a)
            t = buckets[a][ia].get(type_field)

            # 在 b 里挑相同类型且未锁定样本
            cand_b = [j for j, r in enumerate(buckets[b])
                      if r.get(type_field) == t
                      and r["id"] not in locked_by_lane.get(b + 1, set())]
            if not cand_b:
                continue
            ib = random.choice(cand_b)
        else:
            ia = random.randrange(len(buckets[a]))
            ib = random.randrange(len(buckets[b]))
            # 避免换到锁定样本
            if (buckets[a][ia]["id"] in locked_by_lane.get(a + 1, set())
                or buckets[b][ib]["id"] in locked_by_lane.get(b + 1, set())):
                continue

        ra = buckets[a][ia]
        rb = buckets[b][ib]

        # 尝试交换
        buckets[a][ia], buckets[b][ib] = rb, ra

        new_a = cost(buckets[a])
        new_b = cost(buckets[b])
        new_total = total - lane_scores[a] - lane_scores[b] + new_a + new_b

        if new_total < total:
            lane_scores[a], lane_scores[b] = new_a, new_b
            total = new_total
        else:
            # 回滚
            buckets[a][ia], buckets[b][ib] = ra, rb

    return buckets

# -------------------- Pretty printing --------------------

def fmt_bool(b: bool) -> str:
    return "PASS" if b else "FAIL"

def fmt_color_triplet(d: Dict[str, float]) -> str:
    return f"green={d['green']:.2f}  blue={d['blue']:.2f}  dark={d['dark']:.2f}"

def print_per_cycle_list(per_cycle: List[Dict[str, Any]], indent: str = "  ") -> None:
    print(f"{indent}Per-cycle results:")
    for pc in per_cycle:
        print(f"{indent}  #{pc['cycle']:>2}  {fmt_color_triplet(pc['color'])}  [{'PASS' if pc['ok'] else 'FAIL'}]")

def print_check_human(tag: str, rep: Dict[str, Any], min_green: float, min_blue: float, max_dark: float) -> None:
    print(f"{tag}: {fmt_bool(rep['hard_ok'])}  (HARD report: cycles 1-2 must pass)")
    # c1/c2 details
    c1, c2 = rep["hard_c12"]
    if c1 is not None:
        ok1 = (c1["green"] >= min_green) and (c1["blue"] > min_blue) and (c1["dark"] <= max_dark)
        print(f"  C1: {fmt_color_triplet(c1)}  [{'PASS' if ok1 else 'FAIL'}]")
    else:
        print(f"  C1: N/A")
    if c2 is not None:
        ok2 = (c2["green"] >= min_green) and (c2["blue"] > min_blue) and (c2["dark"] <= max_dark)
        print(f"  C2: {fmt_color_triplet(c2)}  [{'PASS' if ok2 else 'FAIL'}]")
    else:
        print(f"  C2: N/A")

    print(f"  Mean color (1..{rep['cycles']}): {fmt_color_triplet(rep['mean_color'])}")
    wc = rep['worst_cycle']
    if wc:
        print(f"  Worst cycle: #{wc['cycle']}  {fmt_color_triplet(wc['color'])}  [{'PASS' if wc['color_ok'] else 'FAIL'}]")
    # weighted penalty info
    weights_info = rep["weights"]
    if any(w > 0 for w in weights_info):
        print(f"  Weighted penalty (cycles ≥3): {rep['weighted_penalty']:.4f}  "
              f"(w34={sum(weights_info[2:4]):.2f} over C3-4)")
    # per-cycle dump
    print_per_cycle_list(rep["per_cycle"], indent="  ")
    print()

def print_group_human(report: Dict[str, Any], out_path: str = "") -> None:
    print("=== Lane Assignment (Color Balance) ===")
    total_libs = sum(l['n_libraries'] for l in report['lane_reports'])
    print(f"Total libraries: {total_libs}")
    print(f"Cycles used: i7={report['cycles']['i7']}, i5={report['cycles']['i5']}")
    print(f"Number of lanes: {report['lanes']}\n")

    for lane in report["lane_reports"]:
        status = "PASS" if (lane["i7"]["hard_ok"] and lane["i5"]["hard_ok"]) else "FAIL"
        print(f"Lane {lane['lane']}: {status}  ({lane['n_libraries']} libraries)")
        print(f"  i7 mean: {fmt_color_triplet(lane['i7']['mean_color'])}")
        print(f"  i5 mean: {fmt_color_triplet(lane['i5']['mean_color'])}")
        print(f"  i7 weighted penalty (≥3): {lane['i7']['weighted_penalty']:.4f}")
        print(f"  i5 weighted penalty (≥3): {lane['i5']['weighted_penalty']:.4f}")
        # Per-cycle details for i7 and i5
        print("  i7:")
        print_per_cycle_list(lane["i7"]["per_cycle"], indent="    ")
        print("  i5:")
        print_per_cycle_list(lane["i5"]["per_cycle"], indent="    ")
        print()

    print(f"All lanes pass (hard C1-2): {'YES' if report['all_lanes_hard_pass'] else 'NO'}")
    if out_path:
        print(f"Assignment saved to: {out_path}")

# -------------------- Commands --------------------


def cmd_check(args):
    items = read_table_check(args.input)
    validate_indexes(items)

    # Determine if lane column is present and used
    def lane_str(x): return str(x).strip()
    has_lane = any(("lane" in r and lane_str(r.get("lane","")) != "") for r in items)

    def evaluate_and_print(sub_items, lanes_label="ALL"):
        # cycles (clip to available for this subset)
        L7_req = args.i7_cycles if args.i7_cycles is not None else args.cycles
        L5_req = args.i5_cycles if args.i5_cycles is not None else args.cycles
        L7 = effective_cycles(sub_items, "i7", L7_req)
        L5 = effective_cycles(sub_items, "i5", L5_req)

        i7 = [r["i7"][:L7] for r in sub_items]
        i5raw = [r["i5"] for r in sub_items]
        i5 = [(revcomp(s) if args.rc_i5 else s)[:L5] for s in i5raw]

        ratio_list = [r["ratio"] for r in sub_items]

        rep7 = evaluate_color_balance_strict_weighted(i7, ratio_list, L7, args.min_green, args.min_blue, args.max_dark, args.w34)
        rep5 = evaluate_color_balance_strict_weighted(i5, ratio_list, L5, args.min_green, args.min_blue, args.max_dark, args.w34)

        report = {
            "lanes": 1,
            "cycles": {"i7": L7, "i5": L5},
            "lane_reports": [{
                "lane": lanes_label ,
                "n_libraries": len(sub_items),
                "i7": rep7,
                "i5": rep5,
                "types": {}
            }],
            "all_lanes_hard_pass": (rep7["hard_ok"] and rep5["hard_ok"])
        }
        print_group_human(report, out_path="")

    if has_lane:
        labels = sorted({str(r.get("lane","")).strip() for r in items if str(r.get("lane","")).strip() != ""})
        for lbl in labels:
            sub = [r for r in items if str(r.get("lane","")).strip() == lbl]
            print(f"=== CHECK for lane {lbl} ===")
            evaluate_and_print(sub, lanes_label=lbl)
            print()
    else:
        print("=== CHECK for all libraries (no lane column) ===")
        evaluate_and_print(items, lanes_label="ALL")


def cmd_group(args):
    items = read_table_group(args.input)
    validate_indexes(items)

    L7_req = args.i7_cycles if args.i7_cycles is not None else args.cycles
    L5_req = args.i5_cycles if args.i5_cycles is not None else args.cycles
    L7 = effective_cycles(items, "i7", L7_req)
    L5 = effective_cycles(items, "i5", L5_req)

    # weights for optimization (≥C3)
    w7 = make_weights_special(L7, args.w34)
    w5 = make_weights_special(L5, args.w34)

    # initial buckets + optimize (feasibility-first on C1/C2 + weighted ≥3)

    # Build initial buckets
    if args.constraints:
        cdata = load_constraints(args.constraints)
        lane_ids, lane_quotas, lane_must = normalize_constraints(cdata)
        if args.lanes and args.lanes != len(lane_ids):
            print(f"[INFO] Overriding --lanes={args.lanes} by constraints specifying {len(lane_ids)} lanes.")
        args.lanes = len(lane_ids)

        check_quota_feasibility(items, lane_ids, lane_quotas)
        buckets = initial_buckets_with_quotas(items, lane_ids, lane_quotas, lane_must)
        locked_by_lane = {lid: set(sids) for lid, sids in lane_must.items()}
        lock_types = True
    else:
        if not args.lanes:
            raise ValueError("--lanes is required when --constraints is not provided.")
        if args.lanes < 1:
            raise ValueError("--lanes must be >= 1")
        buckets = initial_round_robin(items, args.lanes)
        locked_by_lane = {}
        lock_types = False

    buckets = optimize_assignment(
        buckets, L7=L7, L5=L5, rc_i5=args.rc_i5,
        iters=args.iters, seed=args.seed,
        min_green=args.min_green, min_blue=args.min_blue, max_dark=args.max_dark,
        w7=w7, w5=w5,
        c12_weight=args.c12_weight,
        lock_types=lock_types,        # NEW
        type_field="type",            # NEW
        locked_by_lane=locked_by_lane # NEW
    )

    # per-lane reports + optional CSV
    out_rows: List[Dict[str, Any]] = []
    lane_reports: List[Dict[str, Any]] = []
    all_hard_pass = True

    for lane_id, buck in enumerate(buckets, start=1):
        i7 = [r["i7"][:L7] for r in buck]
        i5raw = [r["i5"] for r in buck]
        i5 = [(revcomp(s) if args.rc_i5 else s)[:L5] for s in i5raw]

        ratios = [r["ratio"] for r in buck]

        rep7 = evaluate_color_balance_strict_weighted(i7, ratios, L7, args.min_green, args.min_blue, args.max_dark, args.w34)
        rep5 = evaluate_color_balance_strict_weighted(i5, ratios, L5, args.min_green, args.min_blue, args.max_dark, args.w34)

        # type counts per lane
        type_counts = {}
        for r in buck:
            t = r.get("type")
            if t is not None:
                type_counts[t] = type_counts.get(t, 0) + 1

        lane_reports.append({
            "lane": lane_id,
            "n_libraries": len(buck),
            "i7": rep7,
            "i5": rep5,
            "types": type_counts
        })

        all_hard_pass = all_hard_pass and rep7["hard_ok"] and rep5["hard_ok"]

        for r in buck:
            out_rows.append({"lane": lane_id, "id": r["id"], "i7": r["i7"], "i5": r["i5"]})

    report = {
        "lanes": args.lanes,
        "cycles": {"i7": L7, "i5": L5},
        "lane_reports": lane_reports,
        "all_lanes_hard_pass": all_hard_pass
    }

    print_group_human(report, out_path=args.out)

    # write assignment
    if args.out:
        p = Path(args.out)
        p.parent.mkdir(parents=True, exist_ok=True)

        # 重新读取输入文件，获得原始 id 顺序
        original_items = read_table_group(args.input)
        id_order = [r["id"] for r in original_items]

        # 建立 id -> (lane, i7, i5) 映射
        lane_map = {r["id"]: (lane_id, r["i7"], r["i5"], r["ratio"], r["type"])
                    for lane_id, buck in enumerate(buckets, start=1)
                    for r in buck}

        with p.open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["id", "i7", "i5", "ratio", "lane", "type"])
            w.writeheader()
            for id_ in id_order:  # 按输入文件的顺序写出
                lane, i7, i5, ratio, type = lane_map[id_]
                w.writerow({"id": id_, "i7": i7, "i5": i5, "ratio": ratio, "lane": lane, "type": type})

# -------------------- CLI --------------------

def build_parser():
    p = argparse.ArgumentParser(
        description="Index color-balance checker and lane balancer (2-color; feasibility-first on C1-2; weighted ≥3; per-cycle outputs)."
    )
    sub = p.add_subparsers(dest="cmd", required=True)

    def add_common(a):
        a.add_argument("input", help="Input (GROUP: id,i7,i5,ratio,type | CHECK: id,i7,i5,ratio,lane) in CSV/TSV/XLSX")
        # cycles control
        a.add_argument("--cycles", type=int, default=8,
                       help="Number of cycles to evaluate for BOTH i7 and i5 (default 8)")
        a.add_argument("--i7-cycles", type=int, default=None,
                       help="Override number of cycles for i7 (default: use --cycles)")
        a.add_argument("--i5-cycles", type=int, default=None,
                       help="Override number of cycles for i5 (default: use --cycles)")
        # color thresholds
        a.add_argument("--min-green", type=float, default=0.25,
                       help="Min green fraction per cycle (default 0.25)")
        a.add_argument("--min-blue", type=float, default=0.05,
                       help="Min blue fraction per cycle (default 0.05; hard check uses '> min_blue')")
        a.add_argument("--max-dark", type=float, default=0.40,
                       help="Max dark fraction per cycle (default 0.40)")
        # weighting for ≥C3
        a.add_argument("--w34", type=float, default=0.60,
                       help="Total weight assigned to cycles 3/4 (even split). Remaining weight goes to 5..N (default 0.60).")
        # orientation
        a.add_argument("--rc-i5", action="store_true",
                       help="Reverse-complement i5 before evaluating (default: OFF)")

    # check
    pc = sub.add_parser("check", help="Check ratio-weighted color balance with hard C1-2 (report) and weighted ≥3; print per-cycle results. Input: id,i7,i5,ratio,lane")
    add_common(pc)
    pc.set_defaults(func=cmd_check)

    # group
    pg = sub.add_parser("group", help="Assign libraries into lanes with ratio-weighted color balance (feasibility-first on C1-2 + weighted ≥3); print per-cycle results per lane. Input: id,i7,i5,ratio,type")
    add_common(pg)
    pg.add_argument("--lanes", type=int, required=False, default=0, help="Number of lanes (optional if --constraints provided)")
    pg.add_argument("--constraints", type=str, default="",
                    help="JSON/YAML file specifying per-lane quotas by sample type and must-include samples. If provided, --lanes can be omitted.")
    pg.add_argument("--iters", type=int, default=20000,
                    help="Swap iterations for optimization (default 20000)")
    pg.add_argument("--seed", type=int, default=42, help="Random seed (default 42)")
    pg.add_argument("--c12-weight", type=float, default=1e6,
                    help="Priority weight for cycles 1/2 violations (lexicographic objective). Larger enforces feasibility-first harder. (default 1e6)")
    pg.add_argument("--out", type=str, default="",
                    help="Write lane assignment to CSV (id,i7,i5,lane)")
    pg.set_defaults(func=cmd_group)

    return p

def main():
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)

if __name__ == "__main__":
    main()